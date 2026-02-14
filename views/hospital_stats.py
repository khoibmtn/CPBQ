"""
views/hospital_stats.py - Trang Số liệu toàn viện
=====================================================
Ma trận so sánh thời điểm: Nội trú | Ngoại trú | Tổng
× Chung / Số tiền / Bình quân.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from bq_helper import run_query
from config import PROJECT_ID, DATASET_ID, VIEW_ID


# ── Data helpers (reused pattern from cost_by_dept) ───────────────────────────

@st.cache_data(ttl=300)
def _get_available_year_months() -> pd.DataFrame:
    """Lấy danh sách (năm, tháng) có dữ liệu."""
    query = f"""
        SELECT DISTINCT nam_qt, thang_qt
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        ORDER BY nam_qt, thang_qt
    """
    return run_query(query)


def _get_years(ym_df: pd.DataFrame) -> list:
    return sorted(ym_df["nam_qt"].unique().tolist(), reverse=True)


def _get_months_for_year(ym_df: pd.DataFrame, year: int) -> list:
    return sorted(ym_df[ym_df["nam_qt"] == year]["thang_qt"].unique().tolist())


def _ym_to_int(year: int, month: int) -> int:
    return year * 100 + month


def _format_period_label(from_y, from_m, to_y, to_m) -> str:
    if from_y == to_y and from_m == to_m:
        return f"Tháng {from_m:02d}.{from_y % 100:02d}"
    return f"{from_m:02d}.{from_y % 100:02d}-{to_m:02d}.{to_y % 100:02d}"


# ── Session state ─────────────────────────────────────────────────────────────

_SS_KEY = "_hs_periods"


def _init_periods():
    if _SS_KEY not in st.session_state:
        st.session_state[_SS_KEY] = [{"id": 1}, {"id": 2}]
    if "_hs_next_id" not in st.session_state:
        st.session_state._hs_next_id = 3


def _add_period():
    st.session_state[_SS_KEY].append({"id": st.session_state._hs_next_id})
    st.session_state._hs_next_id += 1


def _remove_period(period_id: int):
    st.session_state[_SS_KEY] = [
        p for p in st.session_state[_SS_KEY] if p["id"] != period_id
    ]


# ── Period colors ─────────────────────────────────────────────────────────────

PERIOD_COLORS = [
    {"bg": "rgba(37, 99, 235, 0.12)", "border": "#2563eb", "label": "#93c5fd"},
    {"bg": "rgba(234, 88, 12, 0.12)", "border": "#ea580c", "label": "#fdba74"},
    {"bg": "rgba(22, 163, 74, 0.12)", "border": "#16a34a", "label": "#86efac"},
    {"bg": "rgba(147, 51, 234, 0.12)", "border": "#9333ea", "label": "#c4b5fd"},
    {"bg": "rgba(220, 38, 38, 0.12)", "border": "#dc2626", "label": "#fca5a5"},
    {"bg": "rgba(13, 148, 136, 0.12)", "border": "#0d9488", "label": "#5eead4"},
]


def _get_color(idx: int) -> dict:
    return PERIOD_COLORS[idx % len(PERIOD_COLORS)]


# ── BigQuery data ─────────────────────────────────────────────────────────────

# Cost fields used in Số tiền & Bình quân blocks
COST_FIELDS = [
    ("Thuốc",       "t_thuoc"),
    ("Xét nghiệm",  "t_xn"),
    ("CĐHA",         "t_cdha"),
    ("Máu",          "t_mau"),
    ("PTTT",         "t_pttt"),
    ("VTYT",         "t_vtyt"),
    ("Tiền khám",    "t_kham"),
    ("Tiền giường",  "t_giuong"),
    ("Tổng chi",     "t_tongchi"),
    ("BHTT",         "t_bhtt"),
    ("BNTT",         "t_bntt"),
]


@st.cache_data(ttl=300)
def _load_period_data_hospital(from_year: int, from_month: int,
                                to_year: int, to_month: int) -> pd.DataFrame:
    """Tổng hợp theo ml2 (Nội trú / Ngoại trú) cho toàn viện."""
    from_ym = from_year * 100 + from_month
    to_ym = to_year * 100 + to_month

    sum_parts = ", ".join(
        f"SUM(IFNULL({field}, 0)) AS {field}" for _, field in COST_FIELDS
    )
    query = f"""
        SELECT
            ml2,
            COUNT(*) AS so_luot,
            SUM(IFNULL(so_ngay_dtri, 0)) AS so_ngay_dtri,
            {sum_parts}
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        WHERE (nam_qt * 100 + thang_qt) BETWEEN {from_ym} AND {to_ym}
          AND khoa IS NOT NULL
        GROUP BY ml2
    """
    return run_query(query)


# ── Formatting helpers ────────────────────────────────────────────────────────

def _fmt(val, is_count=False) -> str:
    """Format number for display. Dash for 0/None."""
    if val is None or (isinstance(val, float) and pd.isna(val)) or val == 0:
        return "-"
    if is_count:
        return f"{int(val):,}"
    return f"{val:,.0f}"


def _fmt_dec(val) -> str:
    """Format decimal (2 dp). e.g. 7.72"""
    if val is None or (isinstance(val, float) and pd.isna(val)) or val == 0:
        return "-"
    return f"{val:.2f}"


def _fmt_pct_change(first_val, last_val) -> str:
    """Tỷ lệ %: (last/first - 1) × 100%. Colored green/red."""
    if not first_val or first_val == 0 or not last_val:
        return "-"
    pct = (last_val / first_val - 1) * 100
    if pct > 0:
        color, sign = "#4ade80", "+"
    elif pct < 0:
        color, sign = "#f87171", ""
    else:
        color, sign = "#94a3b8", ""
    return f'<span style="color:{color};font-weight:600;">{sign}{pct:.1f}%</span>'


def _pct_change_text(first_val, last_val) -> str:
    """Plain text version of pct change for Excel."""
    if not first_val or first_val == 0 or not last_val:
        return ""
    pct = (last_val / first_val - 1) * 100
    sign = "+" if pct > 0 else ""
    return f"{sign}{pct:.1f}%"


def _fmt_diff(first_val, last_val) -> str:
    """Chênh lệch: last - first. Colored green/red."""
    if first_val is None or last_val is None:
        return "-"
    if (isinstance(first_val, float) and pd.isna(first_val)):
        first_val = 0
    if (isinstance(last_val, float) and pd.isna(last_val)):
        last_val = 0
    diff = last_val - first_val
    if diff == 0:
        return "-"
    if diff > 0:
        color, sign = "#4ade80", "+"
    else:
        color, sign = "#f87171", ""
    # Format based on magnitude
    if isinstance(diff, float) and abs(diff) < 100:
        txt = f"{sign}{diff:.2f}"
    else:
        txt = f"{sign}{diff:,.0f}"
    return f'<span style="color:{color};font-weight:600;">{txt}</span>'


def _diff_raw(first_val, last_val):
    """Raw diff value for Excel."""
    if first_val is None:
        first_val = 0
    if last_val is None:
        last_val = 0
    if isinstance(first_val, float) and pd.isna(first_val):
        first_val = 0
    if isinstance(last_val, float) and pd.isna(last_val):
        last_val = 0
    return last_val - first_val


# ── Shared data access helpers ────────────────────────────────────────────────

def _build_data(periods):
    """Parse period dataframes into list of dicts keyed by ml2."""
    data = []
    for p in periods:
        df = p["data"]
        d = {}
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                d[row["ml2"]] = row.to_dict()
        data.append(d)
    return data


def _get(data, pi, ml2, field):
    row = data[pi].get(ml2)
    if row is None:
        return 0
    v = row.get(field, 0)
    return v if v and not (isinstance(v, float) and pd.isna(v)) else 0


def _get_total(data, pi, field):
    return _get(data, pi, "Nội trú", field) + _get(data, pi, "Ngoại trú", field)


def _bq(amount, so_luot):
    if not so_luot or so_luot == 0 or not amount:
        return 0
    return amount / so_luot


# ── Row definition builders ──────────────────────────────────────────────────

GROUPS = ["Nội trú", "Ngoại trú", "Tổng"]


def _build_all_rows(data, n):
    """Build a structured list of all rows for both HTML table and Excel export.
    Each row = {"label": str, "section": bool, "total_style": bool,
                "values": {group: [raw_val_per_period]}}
    """
    rows = []

    def _add_section(title):
        rows.append({"label": title, "section": True, "total_style": False, "values": {}})

    def _add_row(label, val_fn, total_style=False):
        values = {}
        for g in GROUPS:
            values[g] = [val_fn(pi, g) for pi in range(n)]
        rows.append({"label": label, "section": False, "total_style": total_style, "values": values})

    # ── KHỐI 1 — CHUNG ──
    _add_section("Chung")

    def _so_luot(pi, g):
        if g == "Tổng":
            return _get_total(data, pi, "so_luot")
        return _get(data, pi, g, "so_luot")
    _add_row("Số lượt", _so_luot)

    def _so_ngay(pi, g):
        if g == "Ngoại trú":
            return 0
        return _get(data, pi, "Nội trú", "so_ngay_dtri")
    _add_row("Số ngày ĐT", _so_ngay)

    def _ngay_tb(pi, g):
        if g == "Ngoại trú":
            return 0
        luot = _get(data, pi, "Nội trú", "so_luot")
        ngay = _get(data, pi, "Nội trú", "so_ngay_dtri")
        return ngay / luot if luot else 0
    _add_row("Ngày ĐT TB", _ngay_tb)

    # ── KHỐI 2 — SỐ TIỀN ──
    _add_section("Số tiền")
    for label, field in COST_FIELDS:
        def _make_cost(fld):
            def _fn(pi, g):
                if g == "Tổng":
                    return _get_total(data, pi, fld)
                return _get(data, pi, g, fld)
            return _fn
        _add_row(label, _make_cost(field), total_style=(label == "Tổng chi"))

    # ── KHỐI 3 — BÌNH QUÂN ──
    _add_section("Bình quân")
    for label, field in COST_FIELDS:
        def _make_bq(fld):
            def _fn(pi, g):
                if g == "Tổng":
                    return _bq(_get_total(data, pi, fld), _get_total(data, pi, "so_luot"))
                return _bq(_get(data, pi, g, fld), _get(data, pi, g, "so_luot"))
            return _fn
        _add_row(label, _make_bq(field), total_style=(label == "Tổng chi"))

    return rows


# ── HTML table rendering ─────────────────────────────────────────────────────

def _render_hospital_table(periods: list, show_ratio: bool, show_diff: bool):
    """Render ma trận toàn viện."""

    n = len(periods)
    data = _build_data(periods)
    col_span = n + (1 if show_ratio else 0) + (1 if show_diff else 0)

    # ── CSS ──
    css = """
    <style>
    .hs-table {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Inter', sans-serif;
        font-size: 12px;
        margin-top: 0.5rem;
    }
    .hs-table th, .hs-table td {
        padding: 5px 10px;
        border: 1px solid #475569;
    }
    .hs-table th {
        font-weight: 600;
        text-align: center;
        color: #f8fafc;
    }
    .hs-table td {
        text-align: right;
        color: #f1f5f9;
    }
    .hs-table .h1 { background-color: #1e293b; font-size: 13px; }
    .hs-table .h2 { background-color: #334155; font-size: 11px; }
    .hs-table .lbl {
        text-align: left;
        font-weight: 600;
        white-space: nowrap;
        background-color: #334155;
        color: #f8fafc;
        padding-left: 0.8rem;
    }
    .hs-table .sec-hdr td {
        background-color: #0f172a !important;
        color: #60a5fa !important;
        font-weight: 700;
        font-size: 13px;
    }
    .hs-table .row-even { background-color: #1e293b; }
    .hs-table .row-odd  { background-color: #263548; }
    .hs-table .row-total td {
        background-color: #172033 !important;
        color: #f8fafc !important;
        font-weight: 700;
    }
    </style>
    """

    all_rows = _build_all_rows(data, n)
    n_data_cols = col_span * len(GROUPS)

    html = css + '<table class="hs-table">'

    # ── HEADER ROW 1 ──
    html += "<thead><tr>"
    html += f'<th class="h1" rowspan="2" style="min-width:160px;">Toàn BV</th>'
    for g in GROUPS:
        html += f'<th class="h1" colspan="{col_span}">{g}</th>'
    html += "</tr>"

    # ── HEADER ROW 2 ──
    html += "<tr>"
    for _ in GROUPS:
        for p in periods:
            bg = p["color"]["border"]
            html += f'<th class="h2" style="background-color:{bg};">{p["period_text"]}</th>'
        if show_diff:
            html += '<th class="h2" style="background-color:#d97706;">Chênh lệch</th>'
        if show_ratio:
            html += '<th class="h2" style="background-color:#7c3aed;">Tỷ lệ %</th>'
    html += "</tr></thead>"

    # ── BODY ──
    html += "<tbody>"
    row_idx = 0

    for r in all_rows:
        if r["section"]:
            html += '<tr class="sec-hdr">'
            html += f'<td class="lbl" style="color:#60a5fa !important;background-color:#0f172a !important;">{r["label"]}</td>'
            html += f'<td colspan="{n_data_cols}" style="background-color:#0f172a;"></td>'
            html += "</tr>"
            row_idx = 0
            continue

        is_decimal = r["label"] == "Ngày ĐT TB"
        is_count = r["label"] == "Số lượt"
        cls = "row-total" if r["total_style"] else ("row-even" if row_idx % 2 == 0 else "row-odd")
        row_idx += 1

        html += f'<tr class="{cls}">'
        html += f'<td class="lbl">{r["label"]}</td>'

        for g in GROUPS:
            vals = r["values"][g]
            for pi in range(n):
                v = vals[pi]
                if is_decimal:
                    html += f"<td>{_fmt_dec(v)}</td>"
                elif is_count:
                    html += f"<td>{_fmt(v, is_count=True)}</td>"
                else:
                    html += f"<td>{_fmt(v)}</td>"

            if show_diff:
                first_raw = vals[0]
                last_raw = vals[-1]
                html += f"<td>{_fmt_diff(first_raw, last_raw)}</td>"
            if show_ratio:
                first_raw = vals[0]
                last_raw = vals[-1]
                html += f"<td>{_fmt_pct_change(first_raw, last_raw)}</td>"

        html += "</tr>"

    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


# ── Excel export ──────────────────────────────────────────────────────────────

def _export_to_excel(periods: list, show_ratio: bool, show_diff: bool) -> BytesIO:
    """Export the hospital stats table to Excel with merged headers."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    n = len(periods)
    data = _build_data(periods)
    col_span = n + (1 if show_ratio else 0) + (1 if show_diff else 0)
    all_rows = _build_all_rows(data, n)

    wb = Workbook()
    ws = wb.active
    ws.title = "Toàn viện"

    # Styles
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_header = Font(bold=True, size=11)
    font_section = Font(bold=True, size=11, italic=True)
    font_normal = Font(size=10)
    font_bold = Font(bold=True, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    def _cell(r, c, val, font=font_normal, align=align_right):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.alignment = align
        cell.border = border_all
        return cell

    # ── HEADER ROW 1 ──
    _cell(1, 1, "Toàn BV", font_header, align_center)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    data_col = 2
    for g in GROUPS:
        _cell(1, data_col, g, font_header, align_center)
        end_col = data_col + col_span - 1
        if col_span > 1:
            ws.merge_cells(start_row=1, start_column=data_col, end_row=1, end_column=end_col)
        for c in range(data_col, end_col + 1):
            ws.cell(row=1, column=c).border = border_all
        data_col = end_col + 1

    # ── HEADER ROW 2 ──
    data_col = 2
    for _ in GROUPS:
        for p in periods:
            _cell(2, data_col, p["period_text"], font_header, align_center)
            data_col += 1
        if show_diff:
            _cell(2, data_col, "Chênh lệch", font_header, align_center)
            data_col += 1
        if show_ratio:
            _cell(2, data_col, "Tỷ lệ %", font_header, align_center)
            data_col += 1

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 16
    total_data_cols = col_span * len(GROUPS)
    for c in range(2, 2 + total_data_cols):
        ws.column_dimensions[get_column_letter(c)].width = 15

    # ── BODY ──
    row = 3
    for r in all_rows:
        if r["section"]:
            _cell(row, 1, r["label"], font_section, align_left)
            for c in range(2, 2 + total_data_cols):
                _cell(row, c, "", font_section, align_center)
            row += 1
            continue

        is_decimal = r["label"] == "Ngày ĐT TB"
        font = font_bold if r["total_style"] else font_normal
        _cell(row, 1, r["label"], font, align_left)

        data_col = 2
        for g in GROUPS:
            vals = r["values"][g]
            for pi in range(n):
                v = vals[pi]
                if v == 0:
                    _cell(row, data_col, "", font, align_right)
                elif is_decimal:
                    _cell(row, data_col, round(v, 2), font, align_right)
                else:
                    _cell(row, data_col, round(v) if isinstance(v, float) else v, font, align_right)
                data_col += 1

            if show_diff:
                diff = _diff_raw(vals[0], vals[-1])
                if diff == 0:
                    _cell(row, data_col, "", font, align_right)
                elif is_decimal:
                    _cell(row, data_col, round(diff, 2), font, align_right)
                else:
                    _cell(row, data_col, round(diff) if isinstance(diff, float) else diff, font, align_right)
                data_col += 1

            if show_ratio:
                _cell(row, data_col, _pct_change_text(vals[0], vals[-1]), font, align_center)
                data_col += 1

        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Main render ───────────────────────────────────────────────────────────────

def render():
    """Render trang Số liệu toàn viện."""

    st.markdown("""
    <div class="main-header" style="background: linear-gradient(135deg, #7c3aed, #4f46e5);">
        <h1>🏛️ Số liệu toàn viện</h1>
        <p>Báo cáo hoạt động toàn bệnh viện · So sánh nhiều khoảng thời gian</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Load available year-months ──
    ym_df = _get_available_year_months()
    if ym_df.empty:
        st.warning("⚠️ Chưa có dữ liệu trong database.")
        return

    years = _get_years(ym_df)
    _init_periods()
    periods = st.session_state[_SS_KEY]

    # ── Period section CSS ──
    st.markdown("""
    <style>
        .hs-period-title {
            font-size: 0.9rem;
            font-weight: 700;
            letter-spacing: 0.05em;
            color: #cbd5e1;
        }
        .hs-badge {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 28px;
            height: 28px;
            border-radius: 50%;
            font-size: 0.8rem;
            font-weight: 700;
            color: #fff;
            flex-shrink: 0;
        }
    </style>
    """, unsafe_allow_html=True)

    # ── Header row: title + add button ──
    hdr_left, hdr_right = st.columns([4, 1.5])
    with hdr_left:
        st.markdown('<div class="hs-period-title">KHOẢNG THỜI GIAN SO SÁNH</div>', unsafe_allow_html=True)
    with hdr_right:
        st.button("➕ Thêm khoảng so sánh", key="_hs_add_btn", on_click=_add_period, type="primary")

    # ── Period selectors ──
    collected_periods = []

    for idx, period in enumerate(periods):
        pid = period["id"]
        color = _get_color(idx)
        num = idx + 1

        # Layout: [badge] [from_year] [from_month] [→] [to_year] [to_month] [🗑️]
        cols = st.columns([0.3, 1.2, 1, 0.2, 1.2, 1, 0.3])

        with cols[0]:
            st.markdown(
                f'<div style="padding-top:0.35rem;">'
                f'<span class="hs-badge" style="background-color:{color["border"]};">{num}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        saved_fy = f"_saved_hs_fy_{pid}"
        saved_fm = f"_saved_hs_fm_{pid}"
        saved_ty = f"_saved_hs_ty_{pid}"
        saved_tm = f"_saved_hs_tm_{pid}"

        with cols[1]:
            default_fy_idx = 0
            if saved_fy in st.session_state and st.session_state[saved_fy] in years:
                default_fy_idx = years.index(st.session_state[saved_fy])
            from_year = st.selectbox(
                "Từ năm", years, index=default_fy_idx,
                key=f"_wgt_hs_fy_{pid}", label_visibility="collapsed",
            )
            st.session_state[saved_fy] = from_year

        from_months = _get_months_for_year(ym_df, from_year)
        with cols[2]:
            default_fm_idx = 0
            if saved_fm in st.session_state:
                sv = st.session_state[saved_fm]
                if sv in from_months:
                    default_fm_idx = from_months.index(sv)
            from_month = st.selectbox(
                "Từ tháng", from_months, index=default_fm_idx,
                key=f"_wgt_hs_fm_{pid}", format_func=lambda m: f"Tháng {m:02d}",
                label_visibility="collapsed",
            )
            st.session_state[saved_fm] = from_month

        with cols[3]:
            st.markdown(
                '<div style="text-align:center;padding-top:0.6rem;color:#94a3b8;">→</div>',
                unsafe_allow_html=True,
            )

        with cols[4]:
            default_ty_idx = 0
            if saved_ty in st.session_state and st.session_state[saved_ty] in years:
                default_ty_idx = years.index(st.session_state[saved_ty])
            to_year = st.selectbox(
                "Đến năm", years, index=default_ty_idx,
                key=f"_wgt_hs_ty_{pid}", label_visibility="collapsed",
            )
            st.session_state[saved_ty] = to_year

        to_months = _get_months_for_year(ym_df, to_year)
        with cols[5]:
            from_ym = _ym_to_int(from_year, from_month)
            valid_to_months = [m for m in to_months if _ym_to_int(to_year, m) >= from_ym]
            if not valid_to_months:
                valid_to_months = to_months

            default_tm_idx = len(valid_to_months) - 1
            if saved_tm in st.session_state:
                sv = st.session_state[saved_tm]
                if sv in valid_to_months:
                    default_tm_idx = valid_to_months.index(sv)

            to_month = st.selectbox(
                "Đến tháng", valid_to_months, index=default_tm_idx,
                key=f"_wgt_hs_tm_{pid}", format_func=lambda m: f"Tháng {m:02d}",
                label_visibility="collapsed",
            )
            st.session_state[saved_tm] = to_month

        with cols[6]:
            if len(periods) > 1:
                if st.button("🗑️", key=f"_rm_hs_{pid}", help="Xóa khoảng thời gian này"):
                    _remove_period(pid)
                    st.rerun()

        label = f"Số liệu {num}"
        final_from = _ym_to_int(from_year, from_month)
        final_to = _ym_to_int(to_year, to_month)
        if final_from > final_to:
            st.warning(f"⚠️ {label}: Thời gian 'Từ' phải trước hoặc bằng thời gian 'Đến'.")
        else:
            collected_periods.append({
                "label": label,
                "from_year": from_year, "from_month": from_month,
                "to_year": to_year, "to_month": to_month,
                "period_text": _format_period_label(from_year, from_month, to_year, to_month),
                "color": color,
            })

    st.markdown("---")

    if not collected_periods:
        st.warning("⚠️ Vui lòng cấu hình ít nhất 1 khoảng thời gian hợp lệ.")
        return

    with st.spinner("⏳ Đang truy vấn dữ liệu..."):
        for p in collected_periods:
            p["data"] = _load_period_data_hospital(
                p["from_year"], p["from_month"],
                p["to_year"], p["to_month"],
            )

    # ── Checkboxes + Download button ──
    can_compare = len(collected_periods) >= 2
    col_ratio, col_diff, col_spacer, col_dl = st.columns([1, 1, 3, 1])

    with col_ratio:
        show_ratio = st.checkbox(
            "Tỷ lệ %",
            key="hs_show_ratio",
            disabled=not can_compare,
            help="Thêm cột Tỷ lệ% so sánh khoảng cuối vs đầu" if can_compare else "Cần ≥ 2 khoảng thời gian",
        )
    with col_diff:
        show_diff = st.checkbox(
            "Chênh lệch",
            key="hs_show_diff",
            disabled=not can_compare,
            help="Thêm cột chênh lệch = giá trị cuối − giá trị đầu" if can_compare else "Cần ≥ 2 khoảng thời gian",
        )
    with col_dl:
        excel_buf = _export_to_excel(collected_periods, show_ratio=show_ratio, show_diff=show_diff)
        file_label = collected_periods[0]["period_text"] if collected_periods else "data"
        st.download_button(
            label="📥 Tải Excel",
            data=excel_buf,
            file_name=f"Toan_vien_{file_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="hs_download_excel",
        )

    _render_hospital_table(collected_periods, show_ratio=show_ratio, show_diff=show_diff)
