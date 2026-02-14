"""
views/icd_analysis.py - Trang Phân tích ICD
=============================================
Thống kê chi phí theo mã bệnh chính (ICD 3 ký tự), phân nhóm Nội/Ngoại trú.
Lọc tích lũy theo ngưỡng tỷ lệ %.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from bq_helper import run_query
from config import PROJECT_ID, DATASET_ID, VIEW_ID


# ── Data helpers ──────────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def _get_available_year_months() -> pd.DataFrame:
    query = f"""
        SELECT DISTINCT nam_qt, thang_qt
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        ORDER BY nam_qt, thang_qt
    """
    return run_query(query)


def _get_years(ym_df: pd.DataFrame) -> list:
    return sorted(ym_df["nam_qt"].unique().tolist(), reverse=True)


def _get_months_for_year(ym_df: pd.DataFrame, year: int) -> list:
    return sorted(ym_df[ym_df["nam_qt"] == year]["thang_qt"].unique().tolist())


def _ym_to_int(year: int, month: int) -> int:
    return year * 100 + month


def _format_period_label(from_y, from_m, to_y, to_m) -> str:
    if from_y == to_y and from_m == to_m:
        return f"{from_m:02d}.{from_y % 100:02d}"
    return f"{from_m:02d}.{from_y % 100:02d}-{to_m:02d}.{to_y % 100:02d}"


# ── Session state helpers ─────────────────────────────────────────────────────

_SS_KEY = "_icd_periods"


def _init_periods():
    if _SS_KEY not in st.session_state:
        st.session_state[_SS_KEY] = [
            {"id": 1},
            {"id": 2},
        ]
    if "_icd_next_id" not in st.session_state:
        st.session_state._icd_next_id = 3


def _add_period():
    st.session_state[_SS_KEY].append({"id": st.session_state._icd_next_id})
    st.session_state._icd_next_id += 1


def _remove_period(period_id: int):
    st.session_state[_SS_KEY] = [
        p for p in st.session_state[_SS_KEY] if p["id"] != period_id
    ]


# ── Colors ────────────────────────────────────────────────────────────────────

PERIOD_COLORS = [
    {"bg": "rgba(37, 99, 235, 0.12)", "border": "#2563eb", "label": "#93c5fd"},
    {"bg": "rgba(234, 88, 12, 0.12)", "border": "#ea580c", "label": "#fdba74"},
    {"bg": "rgba(22, 163, 74, 0.12)", "border": "#16a34a", "label": "#86efac"},
    {"bg": "rgba(147, 51, 234, 0.12)", "border": "#9333ea", "label": "#c4b5fd"},
    {"bg": "rgba(220, 38, 38, 0.12)", "border": "#dc2626", "label": "#fca5a5"},
    {"bg": "rgba(13, 148, 136, 0.12)", "border": "#0d9488", "label": "#5eead4"},
]


def _get_color(idx: int) -> dict:
    return PERIOD_COLORS[idx % len(PERIOD_COLORS)]


# ── BigQuery query ────────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def _load_icd_data(from_year: int, from_month: int,
                   to_year: int, to_month: int, ml2_filter: str,
                   khoa_filter: str = "all") -> pd.DataFrame:
    """Query tổng hợp theo ma_benh_chinh (+ ml2 nếu cần).

    ml2_filter: 'Nội trú', 'Ngoại trú', or 'all' (Toàn BV)
    khoa_filter: tên khoa cụ thể hoặc 'all' (Toàn BV)
    """
    from_ym = from_year * 100 + from_month
    to_ym = to_year * 100 + to_month

    ml2_clause = ""
    if ml2_filter != "all":
        ml2_clause = f"AND ml2 = '{ml2_filter}'"

    khoa_clause = ""
    if khoa_filter != "all":
        safe_khoa = khoa_filter.replace("'", "\\'")
        khoa_clause = f"AND khoa = '{safe_khoa}'"

    query = f"""
        SELECT
            ma_benh_chinh,
            COUNT(*) AS so_luot,
            SUM(IFNULL(so_ngay_dtri, 0)) AS so_ngay_dtri,
            SUM(IFNULL(t_tongchi, 0)) AS t_tongchi,
            SUM(IFNULL(t_bhtt, 0)) AS t_bhtt
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        WHERE (nam_qt * 100 + thang_qt) BETWEEN {from_ym} AND {to_ym}
          AND ma_benh_chinh IS NOT NULL
          {ml2_clause}
          {khoa_clause}
        GROUP BY ma_benh_chinh
        ORDER BY t_tongchi DESC
    """
    return run_query(query)


@st.cache_data(ttl=300)
def _get_available_khoa(periods_key: str, ml2_filter: str) -> list:
    """Query danh sách khoa có dữ liệu trong các khoảng thời gian đã chọn.

    periods_key: chuỗi đại diện cho tất cả các khoảng thời gian (dùng làm cache key)
    """
    # Parse periods_key back into range conditions
    # Format: "from_ym1-to_ym1,from_ym2-to_ym2,..."
    range_parts = periods_key.split(",")
    or_clauses = []
    for part in range_parts:
        from_ym, to_ym = part.split("-")
        or_clauses.append(f"(nam_qt * 100 + thang_qt) BETWEEN {from_ym} AND {to_ym}")

    where_ranges = " OR ".join(or_clauses)

    ml2_clause = ""
    if ml2_filter != "all":
        ml2_clause = f"AND ml2 = '{ml2_filter}'"

    query = f"""
        SELECT DISTINCT khoa
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        WHERE ({where_ranges})
          AND khoa IS NOT NULL
          {ml2_clause}
        ORDER BY khoa
    """
    df = run_query(query)
    if df is not None and not df.empty:
        return sorted(df["khoa"].tolist())
    return []


# ── Formatting helpers ────────────────────────────────────────────────────────

def _fmt_number(val, decimals=0):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "-"
    if decimals > 0:
        return f"{val:,.{decimals}f}"
    return f"{int(val):,}"


def _fmt_pct(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "-"
    return f"{val:.2f}%"


# ── Table rendering ──────────────────────────────────────────────────────────

def _compute_row_values(row_dict, cost_type, period_total):
    """Compute derived values for a single ICD row."""
    so_luot = row_dict.get("so_luot", 0) or 0
    so_ngay = row_dict.get("so_ngay_dtri", 0) or 0
    tongchi = row_dict.get("t_tongchi", 0) or 0
    bhtt = row_dict.get("t_bhtt", 0) or 0

    ngay_dttb = (so_ngay / so_luot) if so_luot else 0
    if cost_type == "cpbhyt":
        bq_dt = (bhtt / so_luot) if so_luot else 0
    else:
        bq_dt = (tongchi / so_luot) if so_luot else 0

    if cost_type == "soluot":
        total = period_total["so_luot"]
        pct_val = (so_luot / total * 100) if total else 0
    elif cost_type == "tongcp":
        total = period_total["t_tongchi"]
        pct_val = (tongchi / total * 100) if total else 0
    else:
        total = period_total["t_bhtt"]
        pct_val = (bhtt / total * 100) if total else 0

    return {
        "so_luot": so_luot,
        "ngay_dttb": ngay_dttb,
        "bq_dt": bq_dt,
        "pct_val": pct_val,
    }


def _render_icd_table(periods: list, icd_list: list, cost_type: str,
                     pct_col_label: str, diff_metric: str = None,
                     diff_reverse: bool = False):
    """Render bảng thống kê ICD.

    periods: list of dicts with keys: period_text, color, data (DataFrame)
    icd_list: danh sách mã ICD đã lọc tích lũy
    cost_type: 'soluot', 'tongcp', or 'cpbhyt'
    pct_col_label: '%Số lượt', '%Tổng CP', or '%CP BHYT'
    diff_metric: metric key for difference columns (None = no diff)
    """
    n = len(periods)
    show_diff = diff_metric is not None and n >= 2

    # Precompute totals for each period (for % calculation)
    period_totals = []
    for p in periods:
        df = p["data"]
        if df is not None and not df.empty:
            total_so_luot = df["so_luot"].sum()
            total_tongchi = df["t_tongchi"].sum()
            total_bhtt = df["t_bhtt"].sum()
        else:
            total_so_luot = 0
            total_tongchi = 0
            total_bhtt = 0
        period_totals.append({"so_luot": total_so_luot, "t_tongchi": total_tongchi, "t_bhtt": total_bhtt})

    # Diff column labels
    diff_value_label = ""
    if show_diff:
        if diff_metric in ("so_luot",):
            diff_value_label = "Lượt"
        elif diff_metric in ("ngay_dttb",):
            diff_value_label = "Ngày"
        else:
            diff_value_label = "Chi phí"

    # CSS
    css = """
    <style>
    .icd-table {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Inter', sans-serif;
        font-size: 12px;
        margin-top: 0.5rem;
    }
    .icd-table th, .icd-table td {
        padding: 6px 10px;
        border: 1px solid #475569;
    }
    .icd-table th {
        font-weight: 600;
        text-align: center;
        color: #f8fafc;
    }
    .icd-table td {
        text-align: right;
        color: #f1f5f9;
    }
    .icd-table .col-stt {
        text-align: center;
        font-weight: 600;
        width: 40px;
        background-color: #334155;
        color: #cbd5e1;
    }
    .icd-table .col-icd {
        text-align: left;
        font-weight: 600;
        white-space: nowrap;
        background-color: #334155;
        color: #f8fafc;
    }
    .icd-table .grp-header {
        background-color: #1e293b;
        font-weight: 700;
        font-size: 12px;
        letter-spacing: 0.02em;
    }
    .icd-table .sub-header {
        background-color: #334155;
        font-size: 11px;
    }
    .icd-table .row-even { background-color: #1e293b; }
    .icd-table .row-odd  { background-color: #263548; }
    .icd-table .row-total {
        background-color: #172033 !important;
        font-weight: 700;
    }
    .icd-table .row-total td {
        background-color: #172033 !important;
        color: #f8fafc !important;
        font-weight: 700;
    }
    .icd-table .diff-pos { color: #4ade80; }
    .icd-table .diff-neg { color: #f87171; }
    </style>
    """

    # Column layout per period: Số lượt | Ngày ĐTTB | BQĐT | %CP
    bq_label = "BQĐT BHYT" if cost_type == "cpbhyt" else "BQĐT Tổng chi"
    col_labels = ["Số lượt", "Ngày ĐTTB", bq_label, pct_col_label]
    cols_per_period = len(col_labels)

    html = css + '<table class="icd-table">'

    # ── HEADER ROW 1 ──
    html += "<thead>"
    html += "<tr>"
    html += '<th class="grp-header" rowspan="2" style="width:40px;">STT</th>'
    html += '<th class="grp-header" rowspan="2">Mã bệnh</th>'
    for p in periods:
        bg = p["color"]["border"]
        html += f'<th class="grp-header" colspan="{cols_per_period}" style="background-color:{bg};">{p["period_text"]}</th>'
    if show_diff:
        diff_dir_label = "T-P" if diff_reverse else "P-T"
        html += f'<th class="grp-header" colspan="2" style="background-color:#6d28d9;">Chênh lệch ({diff_dir_label})</th>'
    html += "</tr>"

    # ── HEADER ROW 2 ──
    html += "<tr>"
    for p in periods:
        bg = p["color"]["border"]
        for cl in col_labels:
            html += f'<th class="sub-header" style="background-color:{bg};">{cl}</th>'
    if show_diff:
        html += f'<th class="sub-header" style="background-color:#6d28d9;">{diff_value_label}</th>'
        html += '<th class="sub-header" style="background-color:#6d28d9;">%</th>'
    html += "</tr>"
    html += "</thead>"

    # ── BODY ──
    html += "<tbody>"

    # Build lookup: period_idx -> {icd_code -> row_dict}
    period_lookups = []
    for p in periods:
        df = p["data"]
        lookup = {}
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                lookup[row["ma_benh_chinh"]] = row.to_dict()
        period_lookups.append(lookup)

    for stt, icd_code in enumerate(icd_list, 1):
        row_class = "row-even" if stt % 2 == 0 else "row-odd"
        html += f'<tr class="{row_class}">'
        html += f'<td class="col-stt">{stt}</td>'
        html += f'<td class="col-icd">{icd_code}</td>'

        computed_values = []  # store computed values per period for diff
        for pi in range(n):
            row = period_lookups[pi].get(icd_code)
            if row:
                vals = _compute_row_values(row, cost_type, period_totals[pi])
                computed_values.append(vals)

                html += f"<td>{_fmt_number(vals['so_luot'])}</td>"
                html += f"<td>{_fmt_number(vals['ngay_dttb'], 2)}</td>"
                html += f"<td>{_fmt_number(vals['bq_dt'])}</td>"
                html += f"<td>{_fmt_pct(vals['pct_val'])}</td>"
            else:
                computed_values.append(None)
                html += "<td>-</td>" * cols_per_period

        # ── Diff columns ──
        if show_diff:
            first_vals = computed_values[0]
            last_vals = computed_values[-1]
            if first_vals and last_vals:
                if diff_reverse:  # T-P: First - Last
                    val_a = first_vals[diff_metric]
                    val_b = last_vals[diff_metric]
                else:  # P-T: Last - First
                    val_a = last_vals[diff_metric]
                    val_b = first_vals[diff_metric]
                diff_abs = val_a - val_b
                diff_pct = ((diff_abs / val_b) * 100) if val_b else 0

                # Color class
                cls_abs = "diff-pos" if diff_abs > 0 else ("diff-neg" if diff_abs < 0 else "")
                cls_pct = "diff-pos" if diff_pct > 0 else ("diff-neg" if diff_pct < 0 else "")

                # Format based on metric type
                if diff_metric == "so_luot":
                    fmt_abs = _fmt_number(diff_abs)
                elif diff_metric == "ngay_dttb":
                    fmt_abs = _fmt_number(diff_abs, 2)
                elif diff_metric == "pct_val":
                    fmt_abs = _fmt_pct(diff_abs)
                else:
                    fmt_abs = _fmt_number(diff_abs)

                html += f'<td class="{cls_abs}">{fmt_abs}</td>'
                html += f'<td class="{cls_pct}">{_fmt_pct(diff_pct)}</td>'
            else:
                html += "<td>-</td><td>-</td>"

        html += "</tr>"

    # ── Total row (from full dataset, not just filtered icd_list) ──
    html += '<tr class="row-total">'
    html += '<td class="col-stt"></td>'
    html += '<td class="col-icd">TỔNG TOÀN BỘ</td>'

    total_computed = []  # for diff calculation on totals
    for pi in range(n):
        df = periods[pi]["data"]
        if df is not None and not df.empty:
            t_so_luot = df["so_luot"].sum()
            t_so_ngay = df["so_ngay_dtri"].sum()
            t_tongchi = df["t_tongchi"].sum()
            t_bhtt = df["t_bhtt"].sum()

            t_ngay_dttb = (t_so_ngay / t_so_luot) if t_so_luot else 0
            if cost_type == "cpbhyt":
                t_bq_dt = (t_bhtt / t_so_luot) if t_so_luot else 0
            else:
                t_bq_dt = (t_tongchi / t_so_luot) if t_so_luot else 0

            total_computed.append({
                "so_luot": t_so_luot,
                "ngay_dttb": t_ngay_dttb,
                "bq_dt": t_bq_dt,
                "pct_val": 100.0,
            })

            html += f"<td>{_fmt_number(t_so_luot)}</td>"
            html += f"<td>{_fmt_number(t_ngay_dttb, 2)}</td>"
            html += f"<td>{_fmt_number(t_bq_dt)}</td>"
            html += f"<td>{_fmt_pct(100.0)}</td>"
        else:
            total_computed.append(None)
            html += "<td>-</td>" * cols_per_period

    # Diff for total row
    if show_diff:
        first_t = total_computed[0]
        last_t = total_computed[-1]
        if first_t and last_t:
            if diff_reverse:
                val_a = first_t[diff_metric]
                val_b = last_t[diff_metric]
            else:
                val_a = last_t[diff_metric]
                val_b = first_t[diff_metric]
            diff_abs = val_a - val_b
            diff_pct = ((diff_abs / val_b) * 100) if val_b else 0

            cls_abs = "diff-pos" if diff_abs > 0 else ("diff-neg" if diff_abs < 0 else "")
            cls_pct = "diff-pos" if diff_pct > 0 else ("diff-neg" if diff_pct < 0 else "")

            if diff_metric == "so_luot":
                fmt_abs = _fmt_number(diff_abs)
            elif diff_metric == "ngay_dttb":
                fmt_abs = _fmt_number(diff_abs, 2)
            elif diff_metric == "pct_val":
                fmt_abs = _fmt_pct(diff_abs)
            else:
                fmt_abs = _fmt_number(diff_abs)

            html += f'<td class="{cls_abs}">{fmt_abs}</td>'
            html += f'<td class="{cls_pct}">{_fmt_pct(diff_pct)}</td>'
        else:
            html += "<td>-</td><td>-</td>"

    html += "</tr>"

    html += "</tbody></table>"

    st.markdown(html, unsafe_allow_html=True)


# ── Excel export ──────────────────────────────────────────────────────────────

def _export_icd_to_excel(periods: list, icd_list: list, cost_type: str,
                         pct_col_label: str, diff_metric: str = None,
                         diff_reverse: bool = False) -> BytesIO:
    """Generate an Excel workbook matching the ICD table. Returns BytesIO."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    n = len(periods)
    show_diff = diff_metric is not None and n >= 2

    # Precompute totals
    period_totals = []
    for p in periods:
        df = p["data"]
        if df is not None and not df.empty:
            period_totals.append({
                "so_luot": df["so_luot"].sum(),
                "t_tongchi": df["t_tongchi"].sum(),
                "t_bhtt": df["t_bhtt"].sum(),
            })
        else:
            period_totals.append({"so_luot": 0, "t_tongchi": 0, "t_bhtt": 0})

    # Build lookup
    period_lookups = []
    for p in periods:
        df = p["data"]
        lookup = {}
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                lookup[row["ma_benh_chinh"]] = row.to_dict()
        period_lookups.append(lookup)

    bq_label = "BQĐT BHYT" if cost_type == "cpbhyt" else "BQĐT Tổng chi"
    col_labels = ["Số lượt", "Ngày ĐTTB", bq_label, pct_col_label]
    cols_per_period = len(col_labels)

    # Diff label
    if show_diff:
        if diff_metric in ("so_luot",):
            diff_value_label = "Lượt"
        elif diff_metric in ("ngay_dttb",):
            diff_value_label = "Ngày"
        else:
            diff_value_label = "Chi phí"

    # ── Styles ──
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_header = Font(bold=True, size=11, color="000000")
    font_normal = Font(size=11, color="000000")
    font_bold = Font(bold=True, size=11, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Phân tích ICD"

    def _cell(r, c, value, font=font_normal, alignment=align_right):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font
        cell.alignment = alignment
        cell.border = border_all
        return cell

    # ── HEADER ROW 1 ──
    r = 1
    _cell(r, 1, "STT", font_header, align_center)
    _cell(r, 2, "Mã bệnh", font_header, align_center)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=2)

    col = 3
    for p in periods:
        end_col = col + cols_per_period - 1
        _cell(r, col, p["period_text"], font_header, align_center)
        if end_col > col:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=end_col)
        col = end_col + 1

    if show_diff:
        diff_dir_label = "T-P" if diff_reverse else "P-T"
        _cell(r, col, f"Chênh lệch ({diff_dir_label})", font_header, align_center)
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)

    # ── HEADER ROW 2 ──
    r = 2
    col = 3
    for _pi in range(n):
        for cl in col_labels:
            _cell(r, col, cl, font_header, align_center)
            col += 1

    if show_diff:
        _cell(r, col, diff_value_label, font_header, align_center)
        _cell(r, col + 1, "%", font_header, align_center)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12

    # ── DATA ROWS ──
    r = 3
    for stt, icd_code in enumerate(icd_list, 1):
        _cell(r, 1, stt, font_normal, align_center)
        _cell(r, 2, icd_code, font_bold, align_left)

        col = 3
        computed_values = []
        for pi in range(n):
            row = period_lookups[pi].get(icd_code)
            if row:
                vals = _compute_row_values(row, cost_type, period_totals[pi])
                computed_values.append(vals)
                _cell(r, col, vals["so_luot"])
                _cell(r, col + 1, round(vals["ngay_dttb"], 2))
                _cell(r, col + 2, round(vals["bq_dt"]))
                _cell(r, col + 3, round(vals["pct_val"], 2))
            else:
                computed_values.append(None)
                for j in range(cols_per_period):
                    _cell(r, col + j, "")
            col += cols_per_period

        # Diff
        if show_diff:
            first_vals = computed_values[0]
            last_vals = computed_values[-1]
            if first_vals and last_vals:
                if diff_reverse:
                    val_a, val_b = first_vals[diff_metric], last_vals[diff_metric]
                else:
                    val_a, val_b = last_vals[diff_metric], first_vals[diff_metric]
                diff_abs = val_a - val_b
                diff_pct = ((diff_abs / val_b) * 100) if val_b else 0
                _cell(r, col, round(diff_abs, 2))
                _cell(r, col + 1, round(diff_pct, 2))
            else:
                _cell(r, col, "")
                _cell(r, col + 1, "")

        r += 1

    # ── TOTAL ROW ──
    _cell(r, 1, "", font_bold, align_center)
    _cell(r, 2, "TỔNG TOÀN BỘ", font_bold, align_left)

    total_computed = []
    col = 3
    for pi in range(n):
        df = periods[pi]["data"]
        if df is not None and not df.empty:
            t_so_luot = df["so_luot"].sum()
            t_so_ngay = df["so_ngay_dtri"].sum()
            t_tongchi = df["t_tongchi"].sum()
            t_bhtt = df["t_bhtt"].sum()
            t_ngay_dttb = (t_so_ngay / t_so_luot) if t_so_luot else 0
            t_bq_dt = (t_bhtt / t_so_luot) if cost_type == "cpbhyt" else (t_tongchi / t_so_luot) if t_so_luot else 0
            total_computed.append({"so_luot": t_so_luot, "ngay_dttb": t_ngay_dttb, "bq_dt": t_bq_dt, "pct_val": 100.0})
            _cell(r, col, t_so_luot, font_bold)
            _cell(r, col + 1, round(t_ngay_dttb, 2), font_bold)
            _cell(r, col + 2, round(t_bq_dt), font_bold)
            _cell(r, col + 3, 100.0, font_bold)
        else:
            total_computed.append(None)
            for j in range(cols_per_period):
                _cell(r, col + j, "", font_bold)
        col += cols_per_period

    if show_diff:
        first_t = total_computed[0]
        last_t = total_computed[-1]
        if first_t and last_t:
            if diff_reverse:
                val_a, val_b = first_t[diff_metric], last_t[diff_metric]
            else:
                val_a, val_b = last_t[diff_metric], first_t[diff_metric]
            diff_abs = val_a - val_b
            diff_pct = ((diff_abs / val_b) * 100) if val_b else 0
            _cell(r, col, round(diff_abs, 2), font_bold)
            _cell(r, col + 1, round(diff_pct, 2), font_bold)
        else:
            _cell(r, col, "", font_bold)
            _cell(r, col + 1, "", font_bold)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Main render ───────────────────────────────────────────────────────────────

def render():
    st.markdown("""
    <div class="main-header" style="background: linear-gradient(135deg, #7c3aed, #4f46e5);">
        <h1>🔬 Phân tích ICD</h1>
        <p>Thống kê chi phí theo mã bệnh chính · Lọc tích lũy theo tỷ lệ %</p>
    </div>
    """, unsafe_allow_html=True)

    # Load available year-months
    ym_df = _get_available_year_months()
    if ym_df.empty:
        st.warning("⚠️ Chưa có dữ liệu trong database.")
        return

    years = _get_years(ym_df)
    _init_periods()
    periods = st.session_state[_SS_KEY]

    # ── Period section CSS ──
    st.markdown("""
    <style>
        .icd-period-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 0.8rem;
        }
        .icd-period-title {
            font-size: 0.9rem;
            font-weight: 700;
            letter-spacing: 0.05em;
            color: #cbd5e1;
        }
        .icd-badge {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 28px;
            height: 28px;
            border-radius: 50%;
            font-size: 0.8rem;
            font-weight: 700;
            color: #fff;
            flex-shrink: 0;
        }
    </style>
    """, unsafe_allow_html=True)

    # ── Header row: title + add button ──
    hdr_left, hdr_right = st.columns([4, 1.5])
    with hdr_left:
        st.markdown('<div class="icd-period-title">KHOẢNG THỜI GIAN SO SÁNH</div>', unsafe_allow_html=True)
    with hdr_right:
        st.button(
            "➕ Thêm khoảng so sánh",
            key="_icd_add_btn",
            on_click=_add_period,
            type="primary",
        )

    # ── Render each period selector ──
    collected_periods = []

    for idx, period in enumerate(periods):
        pid = period["id"]
        color = _get_color(idx)
        is_required = idx == 0
        num = idx + 1

        # Layout: [badge] [from_year] [from_month] [→] [to_year] [to_month] [🗑️]
        cols = st.columns([0.3, 1.2, 1, 0.2, 1.2, 1, 0.3])

        with cols[0]:
            st.markdown(
                f'<div style="padding-top:0.35rem;">'
                f'<span class="icd-badge" style="background-color:{color["border"]};">{num}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        saved_fy = f"_saved_icd_fy_{pid}"
        saved_fm = f"_saved_icd_fm_{pid}"
        saved_ty = f"_saved_icd_ty_{pid}"
        saved_tm = f"_saved_icd_tm_{pid}"

        with cols[1]:
            default_fy_idx = 0
            if saved_fy in st.session_state and st.session_state[saved_fy] in years:
                default_fy_idx = years.index(st.session_state[saved_fy])
            from_year = st.selectbox(
                "Từ năm", years, index=default_fy_idx,
                key=f"_wgt_icd_fy_{pid}", label_visibility="collapsed",
                help="Năm bắt đầu",
            )
            st.session_state[saved_fy] = from_year

        from_months = _get_months_for_year(ym_df, from_year)
        with cols[2]:
            default_fm_idx = 0
            if saved_fm in st.session_state:
                saved_val = st.session_state[saved_fm]
                if saved_val in from_months:
                    default_fm_idx = from_months.index(saved_val)
            from_month = st.selectbox(
                "Từ tháng", from_months, index=default_fm_idx,
                key=f"_wgt_icd_fm_{pid}", label_visibility="collapsed",
                format_func=lambda m: f"Tháng {m:02d}",
                help="Tháng bắt đầu",
            )
            st.session_state[saved_fm] = from_month

        with cols[3]:
            st.markdown("<div style='text-align:center;padding-top:0.6rem;color:#94a3b8;'>→</div>",
                        unsafe_allow_html=True)

        with cols[4]:
            default_ty_idx = 0
            if saved_ty in st.session_state and st.session_state[saved_ty] in years:
                default_ty_idx = years.index(st.session_state[saved_ty])
            to_year = st.selectbox(
                "Đến năm", years, index=default_ty_idx,
                key=f"_wgt_icd_ty_{pid}", label_visibility="collapsed",
                help="Năm kết thúc",
            )
            st.session_state[saved_ty] = to_year

        to_months = _get_months_for_year(ym_df, to_year)
        with cols[5]:
            default_tm_idx = 0
            if saved_tm in st.session_state:
                saved_val = st.session_state[saved_tm]
                if saved_val in to_months:
                    default_tm_idx = to_months.index(saved_val)
            to_month = st.selectbox(
                "Đến tháng", to_months, index=default_tm_idx,
                key=f"_wgt_icd_tm_{pid}", label_visibility="collapsed",
                format_func=lambda m: f"Tháng {m:02d}",
                help="Tháng kết thúc",
            )
            st.session_state[saved_tm] = to_month

        with cols[6]:
            if len(periods) > 1:
                if st.button("🗑️", key=f"_rm_icd_{pid}", help="Xóa khoảng thời gian này"):
                    _remove_period(pid)
                    st.rerun()

        # Validate
        label = f"Số liệu {num}"
        final_from = _ym_to_int(from_year, from_month)
        final_to = _ym_to_int(to_year, to_month)
        if final_from > final_to:
            st.warning(f"⚠️ {label}: Thời gian 'Từ' phải trước hoặc bằng thời gian 'Đến'.")
        else:
            collected_periods.append({
                "label": label,
                "from_year": from_year, "from_month": from_month,
                "to_year": to_year, "to_month": to_month,
                "period_text": _format_period_label(from_year, from_month, to_year, to_month),
                "color": color,
            })

    st.markdown("---")

    if not collected_periods:
        st.warning("⚠️ Vui lòng cấu hình ít nhất 1 khoảng thời gian hợp lệ.")
        return

    # ── Inline filter bar: all controls in one row ──
    period_text_options = [p["period_text"] for p in collected_periods]
    ml2_options = ["Nội trú", "Ngoại trú", "Toàn BV"]
    ml2_defaults = {"Nội trú": 70, "Ngoại trú": 80, "Toàn BV": 70}

    (col_khoa, col_period, col_cost, col_ml2,
     col_ratio, col_diff_sel, col_diff_toggle) = st.columns(
        [1.5, 0.8, 1.0, 1.0, 0.7, 1.2, 0.5]
    )

    with col_ml2:
        selected_ml2 = st.selectbox(
            "🏥 Loại hình",
            ml2_options,
            key="icd_ml2",
            help="Chọn Nội trú, Ngoại trú hoặc Toàn BV",
        )

    ml2_filter = "all" if selected_ml2 == "Toàn BV" else selected_ml2

    # Build periods key for caching: "from_ym1-to_ym1,from_ym2-to_ym2,..."
    periods_key = ",".join(
        f"{_ym_to_int(p['from_year'], p['from_month'])}-{_ym_to_int(p['to_year'], p['to_month'])}"
        for p in collected_periods
    )
    available_khoa = _get_available_khoa(periods_key, ml2_filter)
    khoa_options = ["Toàn Bệnh viện"] + available_khoa

    with col_khoa:
        selected_khoa = st.selectbox(
            "Thống kê theo khoa",
            khoa_options,
            key="icd_khoa_filter",
            help="Chọn khoa cụ thể hoặc Toàn Bệnh viện",
        )

    khoa_filter = "all" if selected_khoa == "Toàn Bệnh viện" else selected_khoa

    with col_period:
        selected_period_text = st.selectbox(
            "Mốc thống kê",
            period_text_options,
            key="icd_sort_period",
            help="Chọn khoảng thời gian để sắp xếp và lọc tích lũy",
        )

    with col_cost:
        cost_type_label = st.selectbox(
            "Loại thống kê",
            ["Số lượt", "Tổng CP", "CP BHYT"],
            key="icd_cost_type",
            help="Chọn loại thống kê để tính % và lọc tích lũy",
        )

    # Ratio: remember per ml2 type via session state
    ratio_ss_key = f"_icd_ratio_val_{selected_ml2}"
    default_ratio = ml2_defaults.get(selected_ml2, 70)
    if ratio_ss_key not in st.session_state:
        st.session_state[ratio_ss_key] = default_ratio
    with col_ratio:
        ratio = st.number_input(
            "Ngưỡng %",
            min_value=1, max_value=100,
            value=st.session_state[ratio_ss_key],
            key=f"_wgt_icd_ratio_{selected_ml2}",
            help="Ngưỡng tích lũy: cộng dồn % từ trên xuống đến khi ≤ giá trị này",
        )
        st.session_state[ratio_ss_key] = ratio

    # ── Load data ──
    cost_type_map = {"Số lượt": "soluot", "Tổng CP": "tongcp", "CP BHYT": "cpbhyt"}
    cost_type = cost_type_map.get(cost_type_label, "tongcp")
    pct_label_map = {"soluot": "%Số lượt", "tongcp": "%Tổng CP", "cpbhyt": "%CP BHYT"}
    pct_col_label = pct_label_map[cost_type]

    # ── Diff metric + toggle (same inline row) ──
    diff_metric = None
    diff_reverse = False
    has_diff = len(collected_periods) >= 2

    bq_label = "BQĐT BHYT" if cost_type == "cpbhyt" else "BQĐT Tổng chi"
    diff_options_map = {
        "Số lượt": "so_luot",
        "Ngày ĐTTB": "ngay_dttb",
        bq_label: "bq_dt",
        pct_col_label: "pct_val",
    }
    diff_options = ["Không"] + list(diff_options_map.keys())

    with col_diff_sel:
        if has_diff:
            diff_choice = st.selectbox(
                "📊 Chênh lệch theo",
                diff_options,
                key="icd_diff_metric",
                help="Chọn chỉ tiêu để tính chênh lệch giữa các khoảng thời gian",
            )
        else:
            st.selectbox(
                "📊 Chênh lệch theo",
                ["—"],
                key="icd_diff_metric_disabled",
                disabled=True,
                help="Cần ≥ 2 khoảng thời gian để so sánh",
            )
            diff_choice = "Không"

    with col_diff_toggle:
        if has_diff:
            # Read current state to render label above toggle
            current_dir = st.session_state.get("icd_diff_direction", False)
            dir_label = "T-P" if current_dir else "P-T"
            st.markdown(
                f"<div style='font-size:0.82rem;font-weight:400;color:inherit;"
                f"padding-bottom:0.45rem;'>{dir_label}</div>",
                unsafe_allow_html=True,
            )
            diff_reverse = st.toggle(
                "Hướng",
                value=False,
                key="icd_diff_direction",
                help="P-T: Cuối − Đầu · T-P: Đầu − Cuối",
                label_visibility="collapsed",
            )

    if diff_choice != "Không":
        diff_metric = diff_options_map[diff_choice]

    with st.spinner("⏳ Đang truy vấn dữ liệu..."):
        for p in collected_periods:
            p["data"] = _load_icd_data(
                p["from_year"], p["from_month"],
                p["to_year"], p["to_month"],
                ml2_filter,
                khoa_filter,
            )

    # ── Determine ICD list via cumulative filter ──
    sort_period_idx = 0
    for i, p in enumerate(collected_periods):
        if p["period_text"] == selected_period_text:
            sort_period_idx = i
            break

    sort_df = collected_periods[sort_period_idx]["data"]
    if sort_df is None or sort_df.empty:
        st.info("ℹ️ Không có dữ liệu cho khoảng thời gian đã chọn.")
        return

    sort_df = sort_df.copy()
    if cost_type == "soluot":
        total_val = sort_df["so_luot"].sum()
        sort_df["pct"] = sort_df["so_luot"] / total_val * 100 if total_val else 0
    elif cost_type == "tongcp":
        total_val = sort_df["t_tongchi"].sum()
        sort_df["pct"] = sort_df["t_tongchi"] / total_val * 100 if total_val else 0
    else:
        total_val = sort_df["t_bhtt"].sum()
        sort_df["pct"] = sort_df["t_bhtt"] / total_val * 100 if total_val else 0

    sort_df = sort_df.sort_values("pct", ascending=False).reset_index(drop=True)

    cum_sum = 0.0
    icd_list = []
    for _, row in sort_df.iterrows():
        pct = row["pct"]
        if cum_sum + pct > ratio:
            break
        cum_sum += pct
        icd_list.append(row["ma_benh_chinh"])

    if not icd_list:
        if not sort_df.empty:
            icd_list = [sort_df.iloc[0]["ma_benh_chinh"]]

    total_icd_count = len(sort_df)

    col_info, col_dl = st.columns([5, 1])
    with col_info:
        st.markdown(
            f"<div style='color:#94a3b8;font-size:0.85rem;margin-bottom:0.5rem;'>"
            f"📋 Hiển thị <b>{len(icd_list)}</b> / {total_icd_count} mã bệnh "
            f"(tích lũy {pct_col_label} ≈ <b>{cum_sum:.1f}%</b> / {ratio}%)"
            f"</div>",
            unsafe_allow_html=True,
        )
    with col_dl:
        excel_buf = _export_icd_to_excel(
            collected_periods, icd_list, cost_type, pct_col_label,
            diff_metric, diff_reverse,
        )
        file_label = collected_periods[0]["period_text"] if collected_periods else "data"
        st.download_button(
            label="📥 Tải Excel",
            data=excel_buf,
            file_name=f"ICD_{file_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="icd_download_excel",
        )

    # Render table
    _render_icd_table(collected_periods, icd_list, cost_type, pct_col_label,
                      diff_metric, diff_reverse)
