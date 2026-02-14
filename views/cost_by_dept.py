"""
views/cost_by_dept.py - Trang Chi ph√≠ theo khoa
=================================================
So s√°nh d·ªØ li·ªáu gi·ªØa nhi·ªÅu kho·∫£ng th·ªùi gian.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from bq_helper import run_query, get_full_table_id
from config import PROJECT_ID, DATASET_ID, VIEW_ID, LOOKUP_KHOA_TABLE, LOOKUP_CSKCB_TABLE, LOOKUP_PROFILES_TABLE
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ‚îÄ‚îÄ Data helpers ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

@st.cache_data(ttl=300)
def _get_available_year_months() -> pd.DataFrame:
    """L·∫•y danh s√°ch (nƒÉm, th√°ng) c√≥ d·ªØ li·ªáu, s·∫Øp x·∫øp tƒÉng d·∫ßn."""
    query = f"""
        SELECT DISTINCT nam_qt, thang_qt
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        ORDER BY nam_qt, thang_qt
    """
    return run_query(query)


def _get_years(ym_df: pd.DataFrame) -> list:
    """Danh s√°ch nƒÉm (gi·∫£m d·∫ßn) t·ª´ DataFrame year-month."""
    return sorted(ym_df["nam_qt"].unique().tolist(), reverse=True)


def _get_months_for_year(ym_df: pd.DataFrame, year: int) -> list:
    """Danh s√°ch th√°ng c√≥ d·ªØ li·ªáu cho nƒÉm c·ª• th·ªÉ (tƒÉng d·∫ßn)."""
    return sorted(ym_df[ym_df["nam_qt"] == year]["thang_qt"].unique().tolist())


def _ym_to_int(year: int, month: int) -> int:
    """Chuy·ªÉn (nƒÉm, th√°ng) th√†nh s·ªë nguy√™n ƒë·ªÉ so s√°nh: 202501, 202605..."""
    return year * 100 + month


def _format_period_label(from_y, from_m, to_y, to_m) -> str:
    """Chu·ªói ƒë·∫°i di·ªán kho·∫£ng th·ªùi gian: mm.yy-mm.yy ho·∫∑c mm.yy n·∫øu tr√πng."""
    if from_y == to_y and from_m == to_m:
        return f"{from_m:02d}.{from_y % 100:02d}"
    return f"{from_m:02d}.{from_y % 100:02d}-{to_m:02d}.{to_y % 100:02d}"


# ‚îÄ‚îÄ Session state helpers ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

_SS_KEY = "_cbd_periods"  # list of dicts: [{from_year, from_month, to_year, to_month}, ...]


def _init_periods():
    """Kh·ªüi t·∫°o session state cho danh s√°ch kho·∫£ng th·ªùi gian."""
    if _SS_KEY not in st.session_state:
        st.session_state[_SS_KEY] = [
            {"id": 1},  # S·ªë li·ªáu 1 (b·∫Øt bu·ªôc)
            {"id": 2},  # S·ªë li·ªáu 2 (m·∫∑c ƒë·ªãnh c√≥, c√≥ th·ªÉ b·ªè tr·ªëng)
        ]
    if "_cbd_next_id" not in st.session_state:
        st.session_state._cbd_next_id = 3


def _add_period():
    """Th√™m m·ªôt kho·∫£ng th·ªùi gian m·ªõi."""
    st.session_state[_SS_KEY].append({"id": st.session_state._cbd_next_id})
    st.session_state._cbd_next_id += 1


def _remove_period(period_id: int):
    """X√≥a kho·∫£ng th·ªùi gian theo id."""
    st.session_state[_SS_KEY] = [
        p for p in st.session_state[_SS_KEY] if p["id"] != period_id
    ]


# ‚îÄ‚îÄ Colors for each period ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

PERIOD_COLORS = [
    {"bg": "rgba(37, 99, 235, 0.12)", "border": "#2563eb", "label": "#93c5fd"},  # Blue
    {"bg": "rgba(234, 88, 12, 0.12)", "border": "#ea580c", "label": "#fdba74"},  # Orange
    {"bg": "rgba(22, 163, 74, 0.12)", "border": "#16a34a", "label": "#86efac"},  # Green
    {"bg": "rgba(147, 51, 234, 0.12)", "border": "#9333ea", "label": "#c4b5fd"},  # Purple
    {"bg": "rgba(220, 38, 38, 0.12)", "border": "#dc2626", "label": "#fca5a5"},  # Red
    {"bg": "rgba(13, 148, 136, 0.12)", "border": "#0d9488", "label": "#5eead4"},  # Teal
]


def _get_color(idx: int) -> dict:
    return PERIOD_COLORS[idx % len(PERIOD_COLORS)]


# ‚îÄ‚îÄ Main render ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

# Metric definitions: (display_name, db_field, is_count, noi_tru_only)
METRICS = [
    ("S·ªë l∆∞·ª£t KCB",      None,           True,  False),
    ("S·ªë ng√†y ƒëi·ªÅu tr·ªã", "so_ngay_dtri", False, True),   # N·ªôi tr√∫ only
    ("T·ªïng chi",         "t_tongchi",    False, False),
    ("X√©t nghi·ªám",      "t_xn",         False, False),
    ("CƒêHA",             "t_cdha",       False, False),
    ("Thu·ªëc",            "t_thuoc",      False, False),
    ("M√°u",              "t_mau",        False, False),
    ("PTTT",             "t_pttt",       False, False),
    ("VTYT",             "t_vtyt",       False, False),
    ("Ti·ªÅn kh√°m",        "t_kham",       False, False),
    ("Ti·ªÅn gi∆∞·ªùng",      "t_giuong",     False, False),
    ("Ti·ªÅn BHTT",        "t_bhtt",       False, False),
    ("Ti·ªÅn BNTT",        "t_bntt",       False, False),
]

# Metrics that get BQ (b√¨nh qu√¢n / average per visit) columns
# Excludes: S·ªë l∆∞·ª£t KCB (is_count) and S·ªë ng√†y ƒëi·ªÅu tr·ªã
BQ_METRICS = [
    (f"BQ {name}", field, noi_only)
    for name, field, is_count, noi_only in METRICS
    if not is_count and field != "so_ngay_dtri"
]

# Computed ratio metrics: (display_name, numerator_field, denominator_field, noi_only, fmt)
# fmt: "pct" = percentage (√ó100, suffix %), "dec" = decimal (1 decimal place)
RATIO_METRICS = [
    ("Ng√†y ƒêTTB", "so_ngay_dtri", "so_luot", True, "dec"),
    ("T·ª∑ l·ªá thu·ªëc/t·ªïng chi", "t_thuoc", "t_tongchi", False, "pct"),
]


@st.cache_data(ttl=300)
def _load_period_data_by_dept(from_year: int, from_month: int,
                              to_year: int, to_month: int) -> pd.DataFrame:
    """Truy v·∫•n t·ªïng h·ª£p theo khoa + ml2 cho m·ªôt kho·∫£ng th·ªùi gian.

    Returns DataFrame with columns: ml2, khoa, so_luot, so_ngay_dtri,
    t_tongchi, t_xn, t_cdha, t_thuoc, t_mau, t_pttt, t_vtyt, t_kham, t_giuong,
    t_bhtt, t_bntt.
    """
    from_ym = from_year * 100 + from_month
    to_ym = to_year * 100 + to_month

    sum_fields = []
    for _name, field, is_count, noi_only in METRICS:
        if is_count:
            sum_fields.append("COUNT(*) AS so_luot")
        elif noi_only:
            sum_fields.append(f"SUM(IFNULL({field}, 0)) AS {field}")
        else:
            sum_fields.append(f"SUM(IFNULL({field}, 0)) AS {field}")

    select_clause = ",\n            ".join(sum_fields)

    query = f"""
        SELECT
            ml2,
            khoa,
            {select_clause}
        FROM `{PROJECT_ID}.{DATASET_ID}.{VIEW_ID}`
        WHERE (nam_qt * 100 + thang_qt) BETWEEN {from_ym} AND {to_ym}
          AND khoa IS NOT NULL
        GROUP BY ml2, khoa
        ORDER BY ml2, khoa
    """
    return run_query(query)


def _fmt_number(val, is_count=False) -> str:
    """Format s·ªë cho hi·ªÉn th·ªã."""
    if val is None or pd.isna(val) or val == 0:
        return ""
    if is_count:
        return f"{int(val):,}"
    return f"{val:,.0f}"


def _calc_bq(amount, so_luot) -> str:
    """T√≠nh b√¨nh qu√¢n = amount / so_luot, l√†m tr√≤n 0 ch·ªØ s·ªë th·∫≠p ph√¢n."""
    if not so_luot or so_luot == 0:
        return ""
    if amount is None or (isinstance(amount, float) and pd.isna(amount)) or amount == 0:
        return ""
    return f"{round(amount / so_luot):,}"


def _calc_ratio(numerator, denominator, fmt="pct") -> str:
    """T√≠nh t·ª∑ l·ªá = numerator / denominator.
    fmt='pct': hi·ªÉn th·ªã √ó100 + '%';  fmt='dec': hi·ªÉn th·ªã s·ªë th·∫≠p ph√¢n."""
    if not denominator or denominator == 0:
        return ""
    if numerator is None or (isinstance(numerator, float) and pd.isna(numerator)) or numerator == 0:
        return ""
    val = numerator / denominator
    if fmt == "pct":
        return f"{val * 100:.1f}%"
    return f"{val:.1f}"


def _get_col_raw_value(col, pd_row):
    """Extract the raw numeric value from a data row for a given column definition."""
    if pd_row is None:
        return 0
    if col["type"] == "metric":
        key = "so_luot" if col["is_count"] else col["field"]
        return pd_row.get(key, 0) or 0
    elif col["type"] == "bq":
        amount = pd_row.get(col["field"], 0) or 0
        so_luot = pd_row.get("so_luot", 0) or 0
        if not so_luot:
            return 0
        return amount / so_luot
    elif col["type"] == "ratio":
        num = pd_row.get(col["num_field"], 0) or 0
        den = pd_row.get(col["den_field"], 0) or 0
        if not den:
            return 0
        return num / den
    return 0


def _fmt_pct_change(first_val, last_val) -> str:
    """Format percentage change: (last/first - 1) √ó 100%. Green if positive, red if negative."""
    if not first_val or first_val == 0:
        return ""
    if not last_val:
        return ""
    pct = (last_val / first_val - 1) * 100
    if pct > 0:
        color = "#4ade80"  # green
        sign = "+"
    elif pct < 0:
        color = "#f87171"  # red
        sign = ""
    else:
        color = "#94a3b8"  # gray
        sign = ""
    return f'<span style="color:{color};font-weight:600;">{sign}{pct:.1f}%</span>'


def _fmt_diff(first_val, last_val) -> str:
    """Format difference: last - first. Colored green/red."""
    if first_val is None:
        first_val = 0
    if last_val is None:
        last_val = 0
    diff = last_val - first_val
    if diff == 0:
        return "-"
    if diff > 0:
        color, sign = "#4ade80", "+"
    else:
        color, sign = "#f87171", ""
    if isinstance(diff, float) and abs(diff) < 100:
        txt = f"{sign}{diff:.2f}"
    else:
        txt = f"{sign}{diff:,.0f}"
    return f'<span style="color:{color};font-weight:600;">{txt}</span>'


def _diff_raw(first_val, last_val):
    """Raw diff value for Excel."""
    if first_val is None:
        first_val = 0
    if last_val is None:
        last_val = 0
    return last_val - first_val


def _sum_rows(rows: list, n_periods: int) -> list:
    """T√≠nh t·ªïng c√°c d√≤ng d·ªØ li·ªáu cho m·ªói period. Returns list of dicts (one per period)."""
    result = []
    for pi in range(n_periods):
        t = {}
        for _name, field, is_count, _noi_only in METRICS:
            key = "so_luot" if is_count else field
            t[key] = sum(r[pi].get(key, 0) or 0 for r in rows if r[pi])
        result.append(t)
    return result


@st.cache_data(ttl=300)
def _load_khoa_order() -> dict:
    """Load khoa ordering mapping: khoa_name -> thu_tu.

    Handles both direct short_name matches (N·ªôi tr√∫, Th·∫≠n nh√¢n t·∫°o)
    and derived 'Kh√°m b·ªánh (...)' names (Ngo·∫°i tr√∫ Kh√°m b·ªánh).
    'ƒêi·ªÅu tr·ªã ngo·∫°i tr√∫' has no match and will sort last.
    """
    order_query = f"""
        SELECT khoa_name, MIN(thu_tu) AS thu_tu FROM (
            -- N·ªôi tr√∫ + Ngo·∫°i tr√∫ K35: khoa = short_name
            SELECT DISTINCT short_name AS khoa_name, thu_tu
            FROM `{PROJECT_ID}.{DATASET_ID}.{LOOKUP_KHOA_TABLE}`
            WHERE thu_tu IS NOT NULL
            UNION ALL
            -- Ngo·∫°i tr√∫ Kh√°m b·ªánh: khoa = 'Kh√°m b·ªánh (' + ten_cskcb + ')'
            SELECT DISTINCT
                CONCAT('Kh√°m b·ªánh (', IFNULL(cs.ten_cskcb, ''), ')') AS khoa_name,
                kp.thu_tu
            FROM `{PROJECT_ID}.{DATASET_ID}.{LOOKUP_KHOA_TABLE}` kp
            LEFT JOIN `{PROJECT_ID}.{DATASET_ID}.{LOOKUP_CSKCB_TABLE}` cs
                ON CAST(kp.ma_cskcb AS STRING) = CAST(cs.ma_cskcb AS STRING)
            WHERE kp.makhoa_xml = 'K01' AND kp.thu_tu IS NOT NULL
        ) sub
        GROUP BY khoa_name
    """
    try:
        order_df = run_query(order_query)
        return dict(zip(order_df["khoa_name"], order_df["thu_tu"]))
    except Exception:
        return {}


# ‚îÄ‚îÄ Profile helpers ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

# Mapping metric_key ‚Üí (display_name, db_field, is_count, noi_tru_only)
METRIC_LOOKUP = {
    "so_luot":      ("S·ªë l∆∞·ª£t KCB",      None,           True,  False),
    "so_ngay_dtri": ("S·ªë ng√†y ƒëi·ªÅu tr·ªã", "so_ngay_dtri", False, True),
    "t_tongchi":    ("T·ªïng chi",         "t_tongchi",    False, False),
    "t_xn":         ("X√©t nghi·ªám",      "t_xn",         False, False),
    "t_cdha":       ("CƒêHA",             "t_cdha",       False, False),
    "t_thuoc":      ("Thu·ªëc",            "t_thuoc",      False, False),
    "t_mau":        ("M√°u",              "t_mau",        False, False),
    "t_pttt":       ("PTTT",             "t_pttt",       False, False),
    "t_vtyt":       ("VTYT",             "t_vtyt",       False, False),
    "t_kham":       ("Ti·ªÅn kh√°m",        "t_kham",       False, False),
    "t_giuong":     ("Ti·ªÅn gi∆∞·ªùng",      "t_giuong",     False, False),
    "t_bhtt":       ("Ti·ªÅn BHTT",        "t_bhtt",       False, False),
    "t_bntt":       ("Ti·ªÅn BNTT",        "t_bntt",       False, False),
}

# Mapping BQ metric_key ‚Üí (display_name, db_field, noi_tru_only)
BQ_METRIC_LOOKUP = {
    "bq_t_tongchi": ("BQ T·ªïng chi",     "t_tongchi",    False),
    "bq_t_xn":      ("BQ X√©t nghi·ªám",  "t_xn",         False),
    "bq_t_cdha":    ("BQ CƒêHA",         "t_cdha",       False),
    "bq_t_thuoc":   ("BQ Thu·ªëc",        "t_thuoc",      False),
    "bq_t_mau":     ("BQ M√°u",          "t_mau",        False),
    "bq_t_pttt":    ("BQ PTTT",         "t_pttt",       False),
    "bq_t_vtyt":    ("BQ VTYT",         "t_vtyt",       False),
    "bq_t_kham":    ("BQ Ti·ªÅn kh√°m",    "t_kham",       False),
    "bq_t_giuong":  ("BQ Ti·ªÅn gi∆∞·ªùng",  "t_giuong",     False),
    "bq_t_bhtt":    ("BQ BHTT",         "t_bhtt",       False),
    "bq_t_bntt":    ("BQ BNTT",         "t_bntt",       False),
}

# Mapping ratio metric_key ‚Üí (display_name, numerator_field, denominator_field, noi_only, fmt)
RATIO_METRIC_LOOKUP = {
    "ngay_dttb":        ("Ng√†y ƒêTTB",              "so_ngay_dtri", "so_luot",   True,  "dec"),
    "tl_thuoc_tongchi": ("T·ª∑ l·ªá thu·ªëc/t·ªïng chi",   "t_thuoc",      "t_tongchi", False, "pct"),
}


@st.cache_data(ttl=300)
def _load_profile_names_for_page() -> list:
    """Load list of profile names for the dropdown."""
    try:
        full_id = get_full_table_id(LOOKUP_PROFILES_TABLE)
        query = f"SELECT DISTINCT profile_name FROM `{full_id}` ORDER BY profile_name"
        df = run_query(query)
        return df["profile_name"].tolist()
    except Exception:
        return []


@st.cache_data(ttl=300)
def _load_profile_config(profile_name: str) -> list:
    """Load profile config: list of {metric_key, thu_tu, visible} sorted by thu_tu."""
    full_id = get_full_table_id(LOOKUP_PROFILES_TABLE)
    query = f"""
        SELECT metric_key, thu_tu, visible
        FROM `{full_id}`
        WHERE profile_name = '{profile_name}'
        ORDER BY thu_tu
    """
    df = run_query(query)
    return df.to_dict("records")


# Build default unified column list (metrics ‚Üí bq ‚Üí ratio)
DEFAULT_COLUMNS = []
for _name, _field, _is_count, _noi_only in METRICS:
    DEFAULT_COLUMNS.append({"type": "metric", "name": _name, "field": _field, "is_count": _is_count, "noi_only": _noi_only})
for _name, _field, _noi_only in BQ_METRICS:
    DEFAULT_COLUMNS.append({"type": "bq", "name": _name, "field": _field, "noi_only": _noi_only})
for _name, _num, _den, _noi_only, _fmt in RATIO_METRICS:
    DEFAULT_COLUMNS.append({"type": "ratio", "name": _name, "num_field": _num, "den_field": _den, "noi_only": _noi_only, "fmt": _fmt})


def _get_active_columns(profile_name: str) -> list:
    """Get unified ordered column list based on profile.
    Returns list of dicts, each with 'type' ('metric','bq','ratio') and column data.
    Column order respects the profile's thu_tu ordering.
    If profile_name is None (= 'T·∫•t c·∫£'), returns the default list."""
    if profile_name is None:
        return DEFAULT_COLUMNS

    config = _load_profile_config(profile_name)
    if not config:
        return DEFAULT_COLUMNS

    columns = []
    for item in config:
        if not item.get("visible", True):
            continue
        key = item["metric_key"]
        if key in METRIC_LOOKUP:
            name, field, is_count, noi_only = METRIC_LOOKUP[key]
            columns.append({"type": "metric", "name": name, "field": field, "is_count": is_count, "noi_only": noi_only})
        elif key in BQ_METRIC_LOOKUP:
            name, field, noi_only = BQ_METRIC_LOOKUP[key]
            columns.append({"type": "bq", "name": name, "field": field, "noi_only": noi_only})
        elif key in RATIO_METRIC_LOOKUP:
            name, num_field, den_field, noi_only, fmt = RATIO_METRIC_LOOKUP[key]
            columns.append({"type": "ratio", "name": name, "num_field": num_field, "den_field": den_field, "noi_only": noi_only, "fmt": fmt})

    return columns


def _render_comparison_table(periods: list, columns=None, show_ratio=False, show_diff=False):
    """Render b·∫£ng so s√°nh: h√†ng = khoa (nh√≥m Ngo·∫°i tr√∫ ‚Üí N·ªôi tr√∫),
    c·ªôt = nh√≥m ti√™u ch√≠ √ó kho·∫£ng th·ªùi gian."""

    # Use filtered or default columns
    cols = columns if columns is not None else DEFAULT_COLUMNS

    n_periods = len(periods)
    # Only show ratio/diff columns when there are 2+ periods
    show_ratio = show_ratio and n_periods >= 2
    show_diff = show_diff and n_periods >= 2

    # ‚îÄ‚îÄ Load khoa ordering ‚îÄ‚îÄ
    khoa_order = _load_khoa_order()

    # ‚îÄ‚îÄ Shortcut: no columns to show ‚îÄ‚îÄ
    if not cols:
        st.info("Profile n√†y kh√¥ng c√≥ c·ªôt n√†o ƒë∆∞·ª£c hi·ªÉn th·ªã.")
        return

    # ‚îÄ‚îÄ Collect departments grouped by ml2 ‚îÄ‚îÄ
    # Structure: { "Ngo·∫°i tr√∫": {khoa_name: [dict_per_period, ...]}, "N·ªôi tr√∫": ... }
    groups = {"Ngo·∫°i tr√∫": {}, "N·ªôi tr√∫": {}}

    for pi, p in enumerate(periods):
        df = p["data"]
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            ml2 = row["ml2"]
            khoa_name = row["khoa"]
            if ml2 not in groups:
                continue  # skip unknown ml2
            if khoa_name not in groups[ml2]:
                groups[ml2][khoa_name] = [None] * n_periods
            groups[ml2][khoa_name][pi] = row.to_dict()

    # Sort khoa names: by thu_tu ascending, then alphabetical for unordered
    def _sort_key(khoa_name):
        order = khoa_order.get(khoa_name)
        if order is not None:
            return (0, int(order), khoa_name)
        return (1, 0, khoa_name)

    for ml2 in groups:
        groups[ml2] = dict(sorted(groups[ml2].items(), key=lambda x: _sort_key(x[0])))

    # ‚îÄ‚îÄ CSS ‚îÄ‚îÄ
    css = """
    <style>
    .cmp-table {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Inter', sans-serif;
        font-size: 12px;
        margin-top: 0.5rem;
    }
    .cmp-table th, .cmp-table td {
        padding: 6px 10px;
        border: 1px solid #475569;
    }
    .cmp-table th {
        font-weight: 600;
        text-align: center;
        color: #f8fafc;
    }
    .cmp-table td {
        text-align: right;
        color: #f1f5f9;
    }
    .cmp-table .col-tt {
        text-align: center;
        font-weight: 600;
        width: 40px;
        background-color: #334155;
        color: #cbd5e1;
    }
    .cmp-table .col-khoa {
        text-align: left;
        font-weight: 600;
        white-space: nowrap;
        background-color: #334155;
        color: #f8fafc;
    }
    .cmp-table .grp-header {
        background-color: #1e293b;
        font-weight: 700;
        font-size: 12px;
        letter-spacing: 0.02em;
    }
    .cmp-table .sub-header {
        background-color: #334155;
        font-size: 11px;
    }
    .cmp-table .row-even { background-color: #1e293b; }
    .cmp-table .row-odd  { background-color: #263548; }
    .cmp-table .row-section {
        background-color: #0f172a !important;
    }
    .cmp-table .row-section td {
        background-color: #0f172a !important;
        color: #60a5fa !important;
        font-weight: 700;
        font-size: 13px;
    }
    .cmp-table .row-subtotal {
        background-color: #1a2744 !important;
    }
    .cmp-table .row-subtotal td {
        background-color: #1a2744 !important;
        color: #93c5fd !important;
        font-weight: 700;
        font-style: italic;
    }
    .cmp-table .row-total {
        background-color: #172033 !important;
        font-weight: 700;
    }
    .cmp-table .row-total td {
        background-color: #172033 !important;
        color: #f8fafc !important;
        font-weight: 700;
    }
    </style>
    """

    html = css + '<table class="cmp-table">'

    # ‚îÄ‚îÄ HEADER ROW 1: TT | Khoa | column group names (unified order) ‚îÄ‚îÄ
    html += "<thead>"
    html += "<tr>"
    html += '<th class="grp-header" rowspan="2" style="width:40px;">TT</th>'
    html += '<th class="grp-header" rowspan="2">Khoa</th>'
    col_span = n_periods + (1 if show_diff else 0) + (1 if show_ratio else 0)
    for col in cols:
        label = col["name"]
        if col.get("noi_only"):
            label += ' <span style="font-size:10px;opacity:0.7;">(NT)</span>'
        html += f'<th class="grp-header" colspan="{col_span}">{label}</th>'
    html += "</tr>"

    # ‚îÄ‚îÄ HEADER ROW 2: period labels under each column group ‚îÄ‚îÄ
    html += "<tr>"
    for _ in cols:
        for i, p in enumerate(periods):
            bg = p["color"]["border"]
            html += f'<th class="sub-header" style="background-color:{bg};">{p["period_text"]}</th>'
        if show_diff:
            html += '<th class="sub-header" style="background-color:#d97706;">Ch√™nh l·ªách</th>'
        if show_ratio:
            html += '<th class="sub-header" style="background-color:#7c3aed;">T·ª∑ l·ªá%</th>'
    html += "</tr>"
    html += "</thead>"

    # ‚îÄ‚îÄ BODY ‚îÄ‚îÄ
    html += "<tbody>"
    n_data_cols = col_span * len(cols)

    section_labels = [
        ("Ngo·∫°i tr√∫", "I. Ngo·∫°i tr√∫"),
        ("N·ªôi tr√∫",    "II. N·ªôi tr√∫"),
    ]

    all_subtotal_rows = {}  # ml2 -> list of dicts (one per period)

    for ml2_key, section_title in section_labels:
        dept_dict = groups.get(ml2_key, {})
        if not dept_dict:
            continue

        # Section header row
        html += '<tr class="row-section">'
        html += '<td class="col-tt"></td>'
        html += f'<td class="col-khoa" style="color:#60a5fa !important;">{section_title}</td>'
        html += f'<td colspan="{n_data_cols}"></td>'
        html += "</tr>"

        # Department rows
        dept_rows_data = []  # for subtotal calculation
        for idx, (khoa_name, period_data_list) in enumerate(dept_dict.items()):
            row_class = "row-even" if idx % 2 == 0 else "row-odd"
            html += f'<tr class="{row_class}">'
            html += f'<td class="col-tt">{idx + 1}</td>'
            html += f'<td class="col-khoa">{khoa_name}</td>'

            for col in cols:
                for pi in range(n_periods):
                    pd_row = period_data_list[pi]
                    if col["type"] == "metric":
                        key = "so_luot" if col["is_count"] else col["field"]
                        val = pd_row.get(key, 0) if pd_row else 0
                        html += f"<td>{_fmt_number(val, is_count=col['is_count'])}</td>"
                    elif col["type"] == "bq":
                        amount = pd_row.get(col["field"], 0) if pd_row else 0
                        so_luot = pd_row.get("so_luot", 0) if pd_row else 0
                        html += f"<td>{_calc_bq(amount, so_luot)}</td>"
                    elif col["type"] == "ratio":
                        num_val = pd_row.get(col["num_field"], 0) if pd_row else 0
                        den_val = pd_row.get(col["den_field"], 0) if pd_row else 0
                        html += f"<td>{_calc_ratio(num_val, den_val, col['fmt'])}</td>"
                if show_diff:
                    first_val = _get_col_raw_value(col, period_data_list[0])
                    last_val = _get_col_raw_value(col, period_data_list[-1])
                    html += f"<td>{_fmt_diff(first_val, last_val)}</td>"
                if show_ratio:
                    first_val = _get_col_raw_value(col, period_data_list[0])
                    last_val = _get_col_raw_value(col, period_data_list[-1])
                    html += f"<td>{_fmt_pct_change(first_val, last_val)}</td>"

            html += "</tr>"
            dept_rows_data.append(period_data_list)

        # Subtotal for this group
        subtotals = _sum_rows(dept_rows_data, n_periods)
        all_subtotal_rows[ml2_key] = subtotals

    # ‚îÄ‚îÄ III. T·ªîNG section ‚îÄ‚îÄ
    html += '<tr class="row-section">'
    html += '<td class="col-tt"></td>'
    html += f'<td class="col-khoa" style="color:#60a5fa !important;">III. T·ªîNG</td>'
    html += f'<td colspan="{n_data_cols}"></td>'
    html += "</tr>"

    # Row: 1. Ngo·∫°i tr√∫ subtotal
    ngoai_totals = all_subtotal_rows.get("Ngo·∫°i tr√∫", [{} for _ in range(n_periods)])
    html += '<tr class="row-subtotal">'
    html += '<td class="col-tt">1</td>'
    html += '<td class="col-khoa" style="color:#93c5fd !important;">Ngo·∫°i tr√∫</td>'
    for col in cols:
        for pi in range(n_periods):
            if col["type"] == "metric":
                key = "so_luot" if col["is_count"] else col["field"]
                val = ngoai_totals[pi].get(key, 0) if pi < len(ngoai_totals) else 0
                html += f"<td>{_fmt_number(val, is_count=col['is_count'])}</td>"
            elif col["type"] == "bq":
                amount = ngoai_totals[pi].get(col["field"], 0) if pi < len(ngoai_totals) else 0
                so_luot = ngoai_totals[pi].get("so_luot", 0) if pi < len(ngoai_totals) else 0
                html += f"<td>{_calc_bq(amount, so_luot)}</td>"
            elif col["type"] == "ratio":
                num_val = ngoai_totals[pi].get(col["num_field"], 0) if pi < len(ngoai_totals) else 0
                den_val = ngoai_totals[pi].get(col["den_field"], 0) if pi < len(ngoai_totals) else 0
                html += f"<td>{_calc_ratio(num_val, den_val, col['fmt'])}</td>"
        if show_diff:
            first_val = _get_col_raw_value(col, ngoai_totals[0] if 0 < len(ngoai_totals) else {})
            last_val = _get_col_raw_value(col, ngoai_totals[-1] if ngoai_totals else {})
            html += f"<td>{_fmt_diff(first_val, last_val)}</td>"
        if show_ratio:
            first_val = _get_col_raw_value(col, ngoai_totals[0] if 0 < len(ngoai_totals) else {})
            last_val = _get_col_raw_value(col, ngoai_totals[-1] if ngoai_totals else {})
            html += f"<td>{_fmt_pct_change(first_val, last_val)}</td>"
    html += "</tr>"

    # Row: 2. N·ªôi tr√∫ subtotal
    noi_totals = all_subtotal_rows.get("N·ªôi tr√∫", [{} for _ in range(n_periods)])
    html += '<tr class="row-subtotal">'
    html += '<td class="col-tt">2</td>'
    html += '<td class="col-khoa" style="color:#93c5fd !important;">N·ªôi tr√∫</td>'
    for col in cols:
        for pi in range(n_periods):
            if col["type"] == "metric":
                key = "so_luot" if col["is_count"] else col["field"]
                val = noi_totals[pi].get(key, 0) if pi < len(noi_totals) else 0
                html += f"<td>{_fmt_number(val, is_count=col['is_count'])}</td>"
            elif col["type"] == "bq":
                amount = noi_totals[pi].get(col["field"], 0) if pi < len(noi_totals) else 0
                so_luot = noi_totals[pi].get("so_luot", 0) if pi < len(noi_totals) else 0
                html += f"<td>{_calc_bq(amount, so_luot)}</td>"
            elif col["type"] == "ratio":
                num_val = noi_totals[pi].get(col["num_field"], 0) if pi < len(noi_totals) else 0
                den_val = noi_totals[pi].get(col["den_field"], 0) if pi < len(noi_totals) else 0
                html += f"<td>{_calc_ratio(num_val, den_val, col['fmt'])}</td>"
        if show_diff:
            first_val = _get_col_raw_value(col, noi_totals[0] if 0 < len(noi_totals) else {})
            last_val = _get_col_raw_value(col, noi_totals[-1] if noi_totals else {})
            html += f"<td>{_fmt_diff(first_val, last_val)}</td>"
        if show_ratio:
            first_val = _get_col_raw_value(col, noi_totals[0] if 0 < len(noi_totals) else {})
            last_val = _get_col_raw_value(col, noi_totals[-1] if noi_totals else {})
            html += f"<td>{_fmt_pct_change(first_val, last_val)}</td>"
    html += "</tr>"

    # Row: Grand total (Ngo·∫°i tr√∫ + N·ªôi tr√∫)
    html += '<tr class="row-total">'
    html += '<td class="col-tt"></td>'
    html += '<td class="col-khoa">T·ªïng c·ªông</td>'
    # Build combined totals for ratio calculation
    combined_totals = []
    for pi in range(n_periods):
        ct = {}
        for k in (ngoai_totals[pi] if pi < len(ngoai_totals) else {}):
            ct[k] = (ngoai_totals[pi].get(k, 0) or 0) + (noi_totals[pi].get(k, 0) or 0) if pi < len(noi_totals) else (ngoai_totals[pi].get(k, 0) or 0)
        for k in (noi_totals[pi] if pi < len(noi_totals) else {}):
            if k not in ct:
                ct[k] = (noi_totals[pi].get(k, 0) or 0)
        combined_totals.append(ct)

    for col in cols:
        for pi in range(n_periods):
            if col["type"] == "metric":
                key = "so_luot" if col["is_count"] else col["field"]
                ngoai_val = ngoai_totals[pi].get(key, 0) if pi < len(ngoai_totals) else 0
                noi_val = noi_totals[pi].get(key, 0) if pi < len(noi_totals) else 0
                val = (ngoai_val or 0) + (noi_val or 0)
                html += f"<td>{_fmt_number(val, is_count=col['is_count'])}</td>"
            elif col["type"] == "bq":
                ngoai_amount = ngoai_totals[pi].get(col["field"], 0) if pi < len(ngoai_totals) else 0
                noi_amount = noi_totals[pi].get(col["field"], 0) if pi < len(noi_totals) else 0
                ngoai_luot = ngoai_totals[pi].get("so_luot", 0) if pi < len(ngoai_totals) else 0
                noi_luot = noi_totals[pi].get("so_luot", 0) if pi < len(noi_totals) else 0
                total_amount = (ngoai_amount or 0) + (noi_amount or 0)
                total_luot = (ngoai_luot or 0) + (noi_luot or 0)
                html += f"<td>{_calc_bq(total_amount, total_luot)}</td>"
            elif col["type"] == "ratio":
                ngoai_num = ngoai_totals[pi].get(col["num_field"], 0) if pi < len(ngoai_totals) else 0
                noi_num = noi_totals[pi].get(col["num_field"], 0) if pi < len(noi_totals) else 0
                ngoai_den = ngoai_totals[pi].get(col["den_field"], 0) if pi < len(ngoai_totals) else 0
                noi_den = noi_totals[pi].get(col["den_field"], 0) if pi < len(noi_totals) else 0
                total_num = (ngoai_num or 0) + (noi_num or 0)
                total_den = (ngoai_den or 0) + (noi_den or 0)
                html += f"<td>{_calc_ratio(total_num, total_den, col['fmt'])}</td>"
        if show_diff:
            first_val = _get_col_raw_value(col, combined_totals[0] if combined_totals else {})
            last_val = _get_col_raw_value(col, combined_totals[-1] if combined_totals else {})
            html += f"<td>{_fmt_diff(first_val, last_val)}</td>"
        if show_ratio:
            first_val = _get_col_raw_value(col, combined_totals[0] if combined_totals else {})
            last_val = _get_col_raw_value(col, combined_totals[-1] if combined_totals else {})
            html += f"<td>{_fmt_pct_change(first_val, last_val)}</td>"
    html += "</tr>"

    html += "</tbody></table>"
    st.markdown(html, unsafe_allow_html=True)


def _export_to_excel(periods: list, columns=None, show_ratio=False, show_diff=False) -> BytesIO:
    """Generate an Excel workbook matching the displayed comparison table.
    Returns a BytesIO buffer containing the .xlsx file."""

    cols = columns if columns is not None else DEFAULT_COLUMNS
    n_periods = len(periods)
    show_ratio = show_ratio and n_periods >= 2
    show_diff = show_diff and n_periods >= 2
    col_span = n_periods + (1 if show_diff else 0) + (1 if show_ratio else 0)

    # ‚îÄ‚îÄ Load khoa ordering & group data (same logic as _render_comparison_table) ‚îÄ‚îÄ
    khoa_order = _load_khoa_order()
    groups = {"Ngo·∫°i tr√∫": {}, "N·ªôi tr√∫": {}}

    for pi, p in enumerate(periods):
        df = p["data"]
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            ml2, khoa_name = row["ml2"], row["khoa"]
            if ml2 not in groups:
                continue
            if khoa_name not in groups[ml2]:
                groups[ml2][khoa_name] = [None] * n_periods
            groups[ml2][khoa_name][pi] = row.to_dict()

    def _sort_key(khoa_name):
        order = khoa_order.get(khoa_name)
        if order is not None:
            return (0, int(order), khoa_name)
        return (1, 0, khoa_name)

    for ml2 in groups:
        groups[ml2] = dict(sorted(groups[ml2].items(), key=lambda x: _sort_key(x[0])))

    # ‚îÄ‚îÄ Style definitions ‚îÄ‚îÄ
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_header = Font(bold=True, size=11, color="000000")
    font_normal = Font(size=11, color="000000")
    font_bold = Font(bold=True, size=11, color="000000")
    font_section = Font(bold=True, size=12, color="000000")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "CP theo khoa"

    # ‚îÄ‚îÄ Helpers ‚îÄ‚îÄ
    def _cell(r, c, value, font=font_normal, alignment=align_right):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font
        cell.alignment = alignment
        cell.border = border_all
        return cell

    def _pct_change_text(first_val, last_val):
        """Return plain text percentage change (no color)."""
        if not first_val or first_val == 0 or not last_val:
            return ""
        pct = (last_val / first_val - 1) * 100
        sign = "+" if pct > 0 else ""
        return f"{sign}{pct:.1f}%"

    def _diff_cell_value(first_val, last_val):
        """Return raw diff value for Excel."""
        f = first_val if first_val else 0
        l = last_val if last_val else 0
        diff = l - f
        return diff if diff != 0 else ""

    def _col_cell_value(col, pd_row):
        """Compute the display value for a column + data row."""
        if pd_row is None:
            return ""
        if col["type"] == "metric":
            key = "so_luot" if col["is_count"] else col["field"]
            val = pd_row.get(key, 0)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return ""
            return int(val) if col["is_count"] else round(val)
        elif col["type"] == "bq":
            amount = pd_row.get(col["field"], 0) or 0
            so_luot = pd_row.get("so_luot", 0) or 0
            if not so_luot or not amount:
                return ""
            return round(amount / so_luot)
        elif col["type"] == "ratio":
            num = pd_row.get(col["num_field"], 0) or 0
            den = pd_row.get(col["den_field"], 0) or 0
            if not den or not num:
                return ""
            val = num / den
            if col.get("fmt") == "pct":
                return f"{val * 100:.1f}%"
            return f"{val:.1f}"
        return ""

    # ‚îÄ‚îÄ HEADER ROW 1 ‚îÄ‚îÄ
    row = 1
    _cell(row, 1, "TT", font_header, align_center)
    _cell(row, 2, "Khoa", font_header, align_center)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    data_col = 3  # starting column for metrics
    for col in cols:
        label = col["name"]
        if col.get("noi_only"):
            label += " (NT)"
        _cell(row, data_col, label, font_header, align_center)
        end_col = data_col + col_span - 1
        if col_span > 1:
            ws.merge_cells(start_row=row, start_column=data_col, end_row=row, end_column=end_col)
        # Apply border to all merged cells
        for c in range(data_col, end_col + 1):
            ws.cell(row=row, column=c).border = border_all
        data_col = end_col + 1

    # ‚îÄ‚îÄ HEADER ROW 2 ‚îÄ‚îÄ
    row = 2
    data_col = 3
    for _ in cols:
        for p in periods:
            _cell(row, data_col, p["period_text"], font_header, align_center)
            data_col += 1
        if show_diff:
            _cell(row, data_col, "Ch√™nh l·ªách", font_header, align_center)
            data_col += 1
        if show_ratio:
            _cell(row, data_col, "T·ª∑ l·ªá%", font_header, align_center)
            data_col += 1

    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 5   # TT
    ws.column_dimensions[get_column_letter(2)].width = 30  # Khoa
    total_data_cols = len(cols) * col_span
    for c in range(3, 3 + total_data_cols):
        ws.column_dimensions[get_column_letter(c)].width = 15

    # ‚îÄ‚îÄ BODY ‚îÄ‚îÄ
    row = 3
    section_labels = [("Ngo·∫°i tr√∫", "I. Ngo·∫°i tr√∫"), ("N·ªôi tr√∫", "II. N·ªôi tr√∫")]
    all_subtotal_rows = {}

    for ml2_key, section_title in section_labels:
        dept_dict = groups.get(ml2_key, {})
        if not dept_dict:
            continue

        # Section header row
        _cell(row, 1, "", font_section, align_center)
        _cell(row, 2, section_title, font_section, align_left)
        for c in range(3, 3 + total_data_cols):
            _cell(row, c, "", font_section, align_center)
        row += 1

        # Department rows
        dept_rows_data = []
        for idx, (khoa_name, period_data_list) in enumerate(dept_dict.items()):
            _cell(row, 1, idx + 1, font_normal, align_center)
            _cell(row, 2, khoa_name, font_normal, align_left)

            data_col = 3
            for col in cols:
                for pi in range(n_periods):
                    pd_row = period_data_list[pi]
                    val = _col_cell_value(col, pd_row)
                    _cell(row, data_col, val, font_normal, align_right)
                    data_col += 1
                if show_diff:
                    first_val = _get_col_raw_value(col, period_data_list[0])
                    last_val = _get_col_raw_value(col, period_data_list[-1])
                    _cell(row, data_col, _diff_cell_value(first_val, last_val), font_normal, align_right)
                    data_col += 1
                if show_ratio:
                    first_val = _get_col_raw_value(col, period_data_list[0])
                    last_val = _get_col_raw_value(col, period_data_list[-1])
                    _cell(row, data_col, _pct_change_text(first_val, last_val), font_normal, align_center)
                    data_col += 1

            row += 1
            dept_rows_data.append(period_data_list)

        subtotals = _sum_rows(dept_rows_data, n_periods)
        all_subtotal_rows[ml2_key] = subtotals

    # ‚îÄ‚îÄ III. T·ªîNG section ‚îÄ‚îÄ
    _cell(row, 1, "", font_section, align_center)
    _cell(row, 2, "III. T·ªîNG", font_section, align_left)
    for c in range(3, 3 + total_data_cols):
        _cell(row, c, "", font_section, align_center)
    row += 1

    # Row: 1. Ngo·∫°i tr√∫ subtotal
    ngoai_totals = all_subtotal_rows.get("Ngo·∫°i tr√∫", [{} for _ in range(n_periods)])
    _cell(row, 1, 1, font_bold, align_center)
    _cell(row, 2, "Ngo·∫°i tr√∫", font_bold, align_left)
    data_col = 3
    for col in cols:
        for pi in range(n_periods):
            val = _col_cell_value(col, ngoai_totals[pi] if pi < len(ngoai_totals) else None)
            _cell(row, data_col, val, font_bold, align_right)
            data_col += 1
        if show_diff:
            first_val = _get_col_raw_value(col, ngoai_totals[0] if 0 < len(ngoai_totals) else {})
            last_val = _get_col_raw_value(col, ngoai_totals[-1] if ngoai_totals else {})
            _cell(row, data_col, _diff_cell_value(first_val, last_val), font_bold, align_right)
            data_col += 1
        if show_ratio:
            first_val = _get_col_raw_value(col, ngoai_totals[0] if 0 < len(ngoai_totals) else {})
            last_val = _get_col_raw_value(col, ngoai_totals[-1] if ngoai_totals else {})
            _cell(row, data_col, _pct_change_text(first_val, last_val), font_bold, align_center)
            data_col += 1
    row += 1

    # Row: 2. N·ªôi tr√∫ subtotal
    noi_totals = all_subtotal_rows.get("N·ªôi tr√∫", [{} for _ in range(n_periods)])
    _cell(row, 1, 2, font_bold, align_center)
    _cell(row, 2, "N·ªôi tr√∫", font_bold, align_left)
    data_col = 3
    for col in cols:
        for pi in range(n_periods):
            val = _col_cell_value(col, noi_totals[pi] if pi < len(noi_totals) else None)
            _cell(row, data_col, val, font_bold, align_right)
            data_col += 1
        if show_diff:
            first_val = _get_col_raw_value(col, noi_totals[0] if 0 < len(noi_totals) else {})
            last_val = _get_col_raw_value(col, noi_totals[-1] if noi_totals else {})
            _cell(row, data_col, _diff_cell_value(first_val, last_val), font_bold, align_right)
            data_col += 1
        if show_ratio:
            first_val = _get_col_raw_value(col, noi_totals[0] if 0 < len(noi_totals) else {})
            last_val = _get_col_raw_value(col, noi_totals[-1] if noi_totals else {})
            _cell(row, data_col, _pct_change_text(first_val, last_val), font_bold, align_center)
            data_col += 1
    row += 1

    # Grand total row
    combined_totals = []
    for pi in range(n_periods):
        ct = {}
        for k in (ngoai_totals[pi] if pi < len(ngoai_totals) else {}):
            ct[k] = (ngoai_totals[pi].get(k, 0) or 0) + (noi_totals[pi].get(k, 0) or 0) if pi < len(noi_totals) else (ngoai_totals[pi].get(k, 0) or 0)
        for k in (noi_totals[pi] if pi < len(noi_totals) else {}):
            if k not in ct:
                ct[k] = (noi_totals[pi].get(k, 0) or 0)
        combined_totals.append(ct)

    _cell(row, 1, "", font_bold, align_center)
    _cell(row, 2, "T·ªïng c·ªông", font_bold, align_left)
    data_col = 3
    for col in cols:
        for pi in range(n_periods):
            if col["type"] == "metric":
                key = "so_luot" if col["is_count"] else col["field"]
                ngoai_val = ngoai_totals[pi].get(key, 0) if pi < len(ngoai_totals) else 0
                noi_val = noi_totals[pi].get(key, 0) if pi < len(noi_totals) else 0
                val = (ngoai_val or 0) + (noi_val or 0)
                _cell(row, data_col, int(val) if col["is_count"] else round(val), font_bold, align_right)
            elif col["type"] == "bq":
                ngoai_amount = ngoai_totals[pi].get(col["field"], 0) if pi < len(ngoai_totals) else 0
                noi_amount = noi_totals[pi].get(col["field"], 0) if pi < len(noi_totals) else 0
                ngoai_luot = ngoai_totals[pi].get("so_luot", 0) if pi < len(ngoai_totals) else 0
                noi_luot = noi_totals[pi].get("so_luot", 0) if pi < len(noi_totals) else 0
                total_amount = (ngoai_amount or 0) + (noi_amount or 0)
                total_luot = (ngoai_luot or 0) + (noi_luot or 0)
                bq_val = round(total_amount / total_luot) if total_luot else ""
                _cell(row, data_col, bq_val, font_bold, align_right)
            elif col["type"] == "ratio":
                ngoai_num = ngoai_totals[pi].get(col["num_field"], 0) if pi < len(ngoai_totals) else 0
                noi_num = noi_totals[pi].get(col["num_field"], 0) if pi < len(noi_totals) else 0
                ngoai_den = ngoai_totals[pi].get(col["den_field"], 0) if pi < len(ngoai_totals) else 0
                noi_den = noi_totals[pi].get(col["den_field"], 0) if pi < len(noi_totals) else 0
                total_num = (ngoai_num or 0) + (noi_num or 0)
                total_den = (ngoai_den or 0) + (noi_den or 0)
                if total_den:
                    ratio_val = total_num / total_den
                    ratio_text = f"{ratio_val * 100:.1f}%" if col.get("fmt") == "pct" else f"{ratio_val:.1f}"
                else:
                    ratio_text = ""
                _cell(row, data_col, ratio_text, font_bold, align_right)
            data_col += 1
        if show_diff:
            first_val = _get_col_raw_value(col, combined_totals[0] if combined_totals else {})
            last_val = _get_col_raw_value(col, combined_totals[-1] if combined_totals else {})
            _cell(row, data_col, _diff_cell_value(first_val, last_val), font_bold, align_right)
            data_col += 1
        if show_ratio:
            first_val = _get_col_raw_value(col, combined_totals[0] if combined_totals else {})
            last_val = _get_col_raw_value(col, combined_totals[-1] if combined_totals else {})
            _cell(row, data_col, _pct_change_text(first_val, last_val), font_bold, align_center)
            data_col += 1

    # ‚îÄ‚îÄ Save to buffer ‚îÄ‚îÄ
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def render():
    """Render trang Chi ph√≠ theo khoa."""

    st.markdown("""
    <div class="main-header" style="background: linear-gradient(135deg, #16a34a, #0d9488);">
        <h1>üè• Chi ph√≠ theo khoa</h1>
        <p>So s√°nh chi ph√≠ gi·ªØa c√°c kho·∫£ng th·ªùi gian theo khoa ¬∑ ph√≤ng</p>
    </div>
    """, unsafe_allow_html=True)

    # Load available year-months
    ym_df = _get_available_year_months()
    if ym_df.empty:
        st.warning("‚ö†Ô∏è Ch∆∞a c√≥ d·ªØ li·ªáu trong database.")
        return

    years = _get_years(ym_df)
    _init_periods()
    periods = st.session_state[_SS_KEY]

    # ‚îÄ‚îÄ Period section CSS ‚îÄ‚îÄ
    st.markdown("""
    <style>
        .cbd-period-title {
            font-size: 0.9rem;
            font-weight: 700;
            letter-spacing: 0.05em;
            color: #cbd5e1;
        }
        .cbd-badge {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 28px;
            height: 28px;
            border-radius: 50%;
            font-size: 0.8rem;
            font-weight: 700;
            color: #fff;
            flex-shrink: 0;
        }
    </style>
    """, unsafe_allow_html=True)

    # ‚îÄ‚îÄ Header row: title + add button ‚îÄ‚îÄ
    hdr_left, hdr_right = st.columns([4, 1.5])
    with hdr_left:
        st.markdown('<div class="cbd-period-title">KHO·∫¢NG TH·ªúI GIAN SO S√ÅNH</div>', unsafe_allow_html=True)
    with hdr_right:
        st.button(
            "‚ûï Th√™m kho·∫£ng so s√°nh",
            key="_cbd_add_btn",
            on_click=_add_period,
            type="primary",
        )

    # ‚îÄ‚îÄ Render each period selector ‚îÄ‚îÄ
    collected_periods = []

    for idx, period in enumerate(periods):
        pid = period["id"]
        color = _get_color(idx)
        num = idx + 1

        # Layout: [badge] [from_year] [from_month] [‚Üí] [to_year] [to_month] [üóëÔ∏è]
        cols = st.columns([0.3, 1.2, 1, 0.2, 1.2, 1, 0.3])

        with cols[0]:
            st.markdown(
                f'<div style="padding-top:0.35rem;">'
                f'<span class="cbd-badge" style="background-color:{color["border"]};">{num}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # ‚îÄ‚îÄ Saved state keys ‚îÄ‚îÄ
        saved_fy = f"_saved_cbd_fy_{pid}"
        saved_fm = f"_saved_cbd_fm_{pid}"
        saved_ty = f"_saved_cbd_ty_{pid}"
        saved_tm = f"_saved_cbd_tm_{pid}"

        # ‚îÄ‚îÄ FROM Year ‚îÄ‚îÄ
        with cols[1]:
            default_fy_idx = 0
            if saved_fy in st.session_state and st.session_state[saved_fy] in years:
                default_fy_idx = years.index(st.session_state[saved_fy])

            from_year = st.selectbox(
                "T·ª´ nƒÉm", years,
                index=default_fy_idx,
                key=f"_wgt_cbd_fy_{pid}",
                label_visibility="collapsed",
                help="NƒÉm b·∫Øt ƒë·∫ßu",
            )
            st.session_state[saved_fy] = from_year

        # ‚îÄ‚îÄ FROM Month ‚îÄ‚îÄ
        from_months = _get_months_for_year(ym_df, from_year)
        with cols[2]:
            default_fm_idx = 0
            if saved_fm in st.session_state:
                saved_val = st.session_state[saved_fm]
                if saved_val in from_months:
                    default_fm_idx = from_months.index(saved_val)

            from_month = st.selectbox(
                "T·ª´ th√°ng", from_months,
                index=default_fm_idx,
                key=f"_wgt_cbd_fm_{pid}",
                format_func=lambda m: f"Th√°ng {m:02d}",
                label_visibility="collapsed",
                help="Th√°ng b·∫Øt ƒë·∫ßu",
            )
            st.session_state[saved_fm] = from_month

        # ‚îÄ‚îÄ Arrow ‚îÄ‚îÄ
        with cols[3]:
            st.markdown(
                '<div style="text-align:center;padding-top:0.6rem;color:#94a3b8;">‚Üí</div>',
                unsafe_allow_html=True,
            )

        # ‚îÄ‚îÄ TO Year ‚îÄ‚îÄ
        with cols[4]:
            default_ty_idx = 0
            if saved_ty in st.session_state and st.session_state[saved_ty] in years:
                default_ty_idx = years.index(st.session_state[saved_ty])

            to_year = st.selectbox(
                "ƒê·∫øn nƒÉm", years,
                index=default_ty_idx,
                key=f"_wgt_cbd_ty_{pid}",
                label_visibility="collapsed",
                help="NƒÉm k·∫øt th√∫c",
            )
            st.session_state[saved_ty] = to_year

        # ‚îÄ‚îÄ TO Month ‚îÄ‚îÄ
        to_months = _get_months_for_year(ym_df, to_year)
        with cols[5]:
            # Filter: "to" must be >= "from" chronologically
            from_ym = _ym_to_int(from_year, from_month)
            valid_to_months = [m for m in to_months if _ym_to_int(to_year, m) >= from_ym]
            if not valid_to_months:
                valid_to_months = to_months  # fallback

            default_tm_idx = len(valid_to_months) - 1  # m·∫∑c ƒë·ªãnh ch·ªçn th√°ng cu·ªëi
            if saved_tm in st.session_state:
                saved_val = st.session_state[saved_tm]
                if saved_val in valid_to_months:
                    default_tm_idx = valid_to_months.index(saved_val)

            to_month = st.selectbox(
                "ƒê·∫øn th√°ng", valid_to_months,
                index=default_tm_idx,
                key=f"_wgt_cbd_tm_{pid}",
                format_func=lambda m: f"Th√°ng {m:02d}",
                label_visibility="collapsed",
                help="Th√°ng k·∫øt th√∫c",
            )
            st.session_state[saved_tm] = to_month

        # ‚îÄ‚îÄ Remove button ‚îÄ‚îÄ
        with cols[6]:
            if len(periods) > 1:
                if st.button("üóëÔ∏è", key=f"_rm_cbd_{pid}", help="X√≥a kho·∫£ng th·ªùi gian n√†y"):
                    _remove_period(pid)
                    st.rerun()

        # Validate chronological order
        label = f"S·ªë li·ªáu {num}"
        final_from = _ym_to_int(from_year, from_month)
        final_to = _ym_to_int(to_year, to_month)
        if final_from > final_to:
            st.warning(f"‚ö†Ô∏è {label}: Th·ªùi gian 'T·ª´' ph·∫£i tr∆∞·ªõc ho·∫∑c b·∫±ng th·ªùi gian 'ƒê·∫øn'.")
        else:
            collected_periods.append({
                "label": label,
                "from_year": from_year,
                "from_month": from_month,
                "to_year": to_year,
                "to_month": to_month,
                "period_text": _format_period_label(from_year, from_month, to_year, to_month),
                "color": color,
            })

    st.markdown("---")

    # ‚îÄ‚îÄ Load and display comparison table ‚îÄ‚îÄ
    if not collected_periods:
        st.warning("‚ö†Ô∏è Vui l√≤ng c·∫•u h√¨nh √≠t nh·∫•t 1 kho·∫£ng th·ªùi gian h·ª£p l·ªá.")
        return

    # Load data for each period
    with st.spinner("‚è≥ ƒêang truy v·∫•n d·ªØ li·ªáu..."):
        for p in collected_periods:
            p["data"] = _load_period_data_by_dept(
                p["from_year"], p["from_month"],
                p["to_year"], p["to_month"],
            )

    # ‚îÄ‚îÄ Profile selector + ratio checkbox + download ‚îÄ‚îÄ
    profile_names = _load_profile_names_for_page()
    profile_options = ["T·∫•t c·∫£"] + profile_names
    col_pf, col_ratio, col_diff, col_dl = st.columns([2, 1, 1, 1])
    with col_pf:
        selected_profile = st.selectbox(
            "üìä Profile hi·ªÉn th·ªã", profile_options,
            key="cost_dept_profile",
            help="Ch·ªçn profile ƒë·ªÉ l·ªçc v√† s·∫Øp x·∫øp c√°c c·ªôt hi·ªÉn th·ªã",
        )
    can_compare = len(collected_periods) >= 2
    with col_ratio:
        show_ratio = st.checkbox(
            "T·ª∑ l·ªá %",
            key="cost_dept_show_ratio",
            disabled=not can_compare,
            help="Th√™m c·ªôt T·ª∑ l·ªá% so s√°nh kho·∫£ng cu·ªëi vs ƒë·∫ßu" if can_compare else "C·∫ßn ‚â• 2 kho·∫£ng th·ªùi gian",
        )
    with col_diff:
        show_diff = st.checkbox(
            "Ch√™nh l·ªách",
            key="cost_dept_show_diff",
            disabled=not can_compare,
            help="Th√™m c·ªôt ch√™nh l·ªách = gi√° tr·ªã cu·ªëi ‚àí gi√° tr·ªã ƒë·∫ßu" if can_compare else "C·∫ßn ‚â• 2 kho·∫£ng th·ªùi gian",
        )

    chosen = None if selected_profile == "T·∫•t c·∫£" else selected_profile
    active_columns = _get_active_columns(chosen)

    with col_dl:
        excel_buf = _export_to_excel(collected_periods, active_columns, show_ratio=show_ratio, show_diff=show_diff)
        file_label = collected_periods[0]["period_text"] if collected_periods else "data"
        st.download_button(
            label="üì• T·∫£i Excel",
            data=excel_buf,
            file_name=f"CP_theo_khoa_{file_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="cost_dept_download_excel",
        )

    # Build and display comparison table
    _render_comparison_table(collected_periods, active_columns, show_ratio=show_ratio, show_diff=show_diff)
